import streamlit as st
import pandas as pd
import banco as db
import psycopg2
from datetime import datetime
import importlib
importlib.reload(db)
import ia_classificador as ia
importlib.reload(ia)
# Configuração global da página do Streamlit
st.set_page_config(page_title="BuscaDados", layout="wide", initial_sidebar_state="expanded")

# Inicializa o arquivo de banco de dados SQLite e cria as tabelas se não existirem
db.inicializar_banco()

# --- ESTILIZAÇÃO CUSTOMIZADA DO MENU LATERAL (CSS INJECT) ---
st.markdown("""
    <style>
        div[data-testid="stSidebarUserContent"] div[role="radiogroup"] label {
            background-color: rgba(248, 249, 250, 0.9) !important;
            padding: 12px 18px !important;
            border-radius: 12px !important;
            margin-bottom: 10px !important;
            border: 1px solid rgba(233, 236, 239, 0.8) !important;
            box-shadow: 0 2px 4px rgba(0,0,0,0.02) !important;
            transition: all 0.15s ease-out !important;
            cursor: pointer !important;
            width: 100% !important;
            display: flex !important;
            align-items: center !important;
        }
        div[data-testid="stSidebarUserContent"] div[role="radiogroup"] label:hover {
            background-color: rgba(255, 255, 255, 1) !important;
            border-color: rgba(255, 75, 75, 0.3) !important;
            transform: translateX(5px) translateY(-1px) !important;
            box-shadow: 4px 4px 10px rgba(255, 75, 75, 0.1) !important;
        }
        div[data-testid="stSidebarUserContent"] div[role="radiogroup"] label[data-checked="true"] {
            background: linear-gradient(135deg, #ff4b4b 0%, #e82c2c 100%) !important;
            border-color: #ff4b4b !important;
            color: white !important;
            font-weight: 600 !important;
            box-shadow: 0 4px 15px rgba(255, 75, 75, 0.4) !important;
            transform: scale(1.02) !important;
        }
        div[data-testid="stSidebarUserContent"] div[role="radiogroup"] label[data-checked="true"]:hover {
            transform: scale(1.03) translateX(3px) !important;
            box-shadow: 0 6px 20px rgba(255, 75, 75, 0.6) !important;
        }
        div[data-testid="stSidebarUserContent"] div[role="radiogroup"] label span[data-testid="stRadioButtonChoiceIndicator"] {
            display: none !important;
        }
        div[data-testid="stSidebarUserContent"] div[role="radiogroup"] label div[data-testid="stWidgetMarkdownInsideRadio"] {
            padding-left: 0px !important;
            margin-left: 0px !important;
            font-size: 1.05rem !important;
            letter-spacing: 0.3px !important;
        }
        
        /* --- CORREÇÕES VISUAIS DA GRADE (GRID) E PAGINAÇÃO --- */
        /* Alinha botões e textos da grade centralizados verticalmente */
        div[data-testid="stHorizontalBlock"] {
            align-items: center !important;
        }
        /* Deixa os botões da tabela mais achatados para diminuir a altura inútil da linha */
        div[data-testid="stHorizontalBlock"] button {
            min-height: 36px !important;
            padding-top: 0px !important;
            padding-bottom: 0px !important;
        }
        /* Estiliza o dropdown de paginação para combinar com os botões brancos */
        div[data-testid="stSelectbox"] > div[data-baseweb="select"] > div {
            min-height: 36px !important;
            border-radius: 8px !important;
            border: 1px solid rgba(49, 51, 63, 0.2) !important;
        }
    </style>
""", unsafe_allow_html=True)

import auth_utils

# --- CONTROLE DE MENSAGENS E ESTADOS ---
if "mensagem_sucesso" not in st.session_state: st.session_state.mensagem_sucesso = None
if "menu_override" not in st.session_state: st.session_state.menu_override = None
if "radio_selecionado" not in st.session_state: st.session_state.radio_selecionado = "🗺️ Manutenção de Municípios"
if "ultimo_menu_acessado" not in st.session_state: st.session_state.ultimo_menu_acessado = None

if "sub_tela_mun" not in st.session_state: st.session_state.sub_tela_mun = "listar"
if "sub_tela_bai" not in st.session_state: st.session_state.sub_tela_bai = "listar"
if "sub_tela_upm" not in st.session_state: st.session_state.sub_tela_upm = "listar"
if "sub_tela_ser" not in st.session_state: st.session_state.sub_tela_ser = "listar"

if "modo_form_mun" not in st.session_state: st.session_state.modo_form_mun = "cadastro"
if "modo_form_bai" not in st.session_state: st.session_state.modo_form_bai = "cadastro"
if "modo_form_upm" not in st.session_state: st.session_state.modo_form_upm = "cadastro"
if "modo_form_ser" not in st.session_state: st.session_state.modo_form_ser = "cadastro"

if "dados_sel_mun" not in st.session_state: st.session_state.dados_sel_mun = {"ID": None, "Municipio": "", "Estado": "MT - Mato Grosso"}
if "dados_sel_bai" not in st.session_state: st.session_state.dados_sel_bai = {"ID": None, "Bairro": "", "Municipio": ""}
if "dados_sel_upm" not in st.session_state: st.session_state.dados_sel_upm = {"ID": None, "UPM": "", "Descricao": "", "Bairro": "", "Municipio": "", "Estado": ""}
if "dados_sel_ser" not in st.session_state: st.session_state.dados_sel_ser = {"ID": None, "Nome": "", "UrlLogin": "", "UrlConsulta": "", "UrlPdf": "", "Login": "", "Senha": "", "DuplaAutenticacao": "Não", "Tipo": "SROP", "Status": "Ativo", "Tempo_Expiracao_Horas": 4, "Layout_ID": 1}

# --- AUTENTICAÇÃO KEYCLOAK (O "PORTEIRO" DO SISTEMA) ---
# Aqui começa a barreira de segurança. Tudo abaixo daqui só roda se a pessoa for autorizada.

# 1. Cria o espaço na memória (sessão) para guardar o "Crachá" do usuário
if "user_info" not in st.session_state:
    st.session_state.user_info = None

# 2. Quando o usuário volta do Keycloak (após digitar a senha), o Keycloak coloca um 'code' na barra de endereços
query_params = st.query_params
if "code" in query_params and not st.session_state.user_info:
    code = query_params["code"]
    
    # 3. Manda esse código secreto de volta pro Keycloak (nos bastidores) em troca do Token de Acesso (Crachá)
    token_response = auth_utils.exchange_code_for_token(code)
    
    if token_response and "access_token" in token_response:
        access_token = token_response["access_token"]
        id_token = token_response.get("id_token")
        
        # 4. Decodifica (abre) o crachá para ver quem é a pessoa e os cargos dela
        decoded = auth_utils.decode_token(access_token)
        user_info = auth_utils.get_user_info(decoded)
        
        if user_info:
            user_info["id_token"] = id_token # Guarda para podermos fazer o Logout depois
            st.session_state.user_info = user_info # Salva o usuário na memória!
            
            # 5. Limpa a barra de endereços (tira o ?code=...) e recarrega a página com o usuário logado
            st.query_params.clear()
            st.rerun()

# 6. Se, depois de tudo, o usuário AINDA NÃO ESTIVER LOGADO (ou seja, acabou de entrar no site)
if not st.session_state.user_info:
    # --- DESIGN PREMIUM DA TELA DE LOGIN (STREAMLIT) ---
    st.markdown("""
        <style>
            /* Esconde a barra lateral, menu superior e rodapé na tela de login para focar 100% no cartão */
            [data-testid="collapsedControl"] { display: none; }
            [data-testid="stHeader"] { display: none; }
            footer { display: none; }
            
            /* Centralização e background */
            .main {
                background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
            }
            
            /* Estilo do Cartão Neumórfico (Efeito de relevo/vidro) */
            .login-card {
                background: rgba(255, 255, 255, 0.9);
                backdrop-filter: blur(10px);
                padding: 50px 40px;
                border-radius: 24px;
                box-shadow: 0 20px 40px rgba(0, 0, 0, 0.1), 0 1px 3px rgba(0,0,0,0.05);
                text-align: center;
                border: 1px solid rgba(255, 255, 255, 0.5);
                margin-top: 5vh;
            }
            
            .login-icon {
                font-size: 60px;
                margin-bottom: -15px;
                animation: float 3s ease-in-out infinite;
            }
            
            .login-title {
                font-size: 38px;
                font-weight: 900;
                color: #1E293B;
                letter-spacing: -1px;
                margin-bottom: 5px;
            }
            
            .login-subtitle {
                color: #64748B;
                font-size: 16px;
                font-weight: 500;
                margin-bottom: 35px;
            }
            
            .login-divider {
                height: 1px;
                background: linear-gradient(90deg, transparent, rgba(255, 75, 75, 0.3), transparent);
                margin: 25px 0;
            }
            
            @keyframes float {
                0% { transform: translateY(0px); }
                50% { transform: translateY(-10px); }
                100% { transform: translateY(0px); }
            }
        </style>
    """, unsafe_allow_html=True)

    # Layout de 3 colunas para forçar o cartão a ficar centralizado
    _, col_center, _ = st.columns([1, 1.5, 1])
    
    with col_center:
        st.markdown("""
            <div class="login-card">
                <div class="login-icon">
                    <svg xmlns="http://www.w3.org/2000/svg" width="68" height="68" viewBox="0 0 24 24" fill="none" stroke="url(#grad1)" stroke-width="1.5" stroke-linecap="round" stroke-linejoin="round">
                        <defs>
                            <linearGradient id="grad1" x1="0%" y1="0%" x2="100%" y2="100%">
                                <stop offset="0%" style="stop-color:#2563EB;stop-opacity:1" />
                                <stop offset="100%" style="stop-color:#6366F1;stop-opacity:1" />
                            </linearGradient>
                        </defs>
                        <rect x="3" y="11" width="18" height="11" rx="2" ry="2"></rect>
                        <path d="M7 11V7a5 5 0 0 1 10 0v4"></path>
                        <circle cx="12" cy="16" r="1" fill="#2563EB"></circle>
                    </svg>
                </div>
                <div class="login-title">BuscaDados</div>
                <div class="login-subtitle">Plataforma Segura de Extração de Dados</div>
                <div class="login-divider"></div>
                <p style="color: #94A3B8; font-size: 14px; margin-bottom: 25px;">
                    Ambiente restrito. O acesso é monitorado e requer autenticação via SSO (Single Sign-On).
                </p>
            </div>
        """, unsafe_allow_html=True)
        
        login_url = auth_utils.get_login_url()
        # st.link_button("🔐 ACESSAR SISTEMA", login_url, type="primary", use_container_width=True)
        st.markdown(f"""
            <a href="{login_url}" target="_self" style="
                display: block;
                width: 100%;
                background-color: #ff4b4b;
                color: white;
                text-align: center;
                padding: 10px 0;
                border-radius: 8px;
                font-size: 16px;
                font-weight: 500;
                text-decoration: none;
                border: 1px solid transparent;
            " onmouseover="this.style.backgroundColor='#ff3333'" onmouseout="this.style.backgroundColor='#ff4b4b'">
                🔐 ACESSAR SISTEMA
            </a>
        """, unsafe_allow_html=True)
        
        st.markdown("<br><p style='text-align: center; color: #94A3B8; font-size: 12px;'>© 2024 BuscaDados. Todos os direitos reservados.</p>", unsafe_allow_html=True)

    # MATA A EXECUÇÃO AQUI! O código do Streamlit para nesta linha.
    st.stop() 

# 7. Se o código chegou até aqui, é porque a pessoa ESTÁ LOGADA!
user = st.session_state.user_info
is_admin = user.get("is_admin", False)

# --- MENU LATERAL ESQUERDO ---
st.sidebar.title("🤖 BuscaDados")
st.sidebar.subheader("Menu Principal")

opcoes_menu = ["🏙️ Manutenção de Municípios", "🏘️ Manutenção de Bairros", "🚔 Manutenção de UPMs", "🌐 Manutenção de Serviços", "📄 Manutenção de Layouts", "📍 Manutenção de Tipo Local", "📝 Manutenção de Prompts", "⚡ Tratar Planilha (Injetar UPM)", "🤖 Robô de Extração (BO)", "⚙️ Configurações"]

# --- CONTROLE DE ACESSO BASEADO EM PAPEL (RBAC) ---
# Se o usuário NÃO for Administrador (ou seja, é um Operador), ele sofre uma restrição.
if not is_admin:
    # Sobrescrevemos a lista de menus para ter APENAS a extração. Todo o resto some!
    opcoes_menu = ["🤖 Robô de Extração (BO)"]
    if st.session_state.radio_selecionado not in opcoes_menu:
        st.session_state.radio_selecionado = "🤖 Robô de Extração (BO)"

# O index do radio deve ser baseado no que está em st.session_state.radio_selecionado, 
# MAS se estivermos numa tela de serviço (override), o menu principal deve ficar "desmarcado"
if st.session_state.menu_override:
    idx_radio_atual = None
else:
    idx_radio_atual = opcoes_menu.index(st.session_state.radio_selecionado) if st.session_state.radio_selecionado in opcoes_menu else 0

radio_val = st.sidebar.radio(
    "Selecione uma opção:",
    opcoes_menu,
    index=idx_radio_atual,
    label_visibility="collapsed"
)

# Se o usuário clicou ativamente no radio (radio_val != None) e o valor for diferente do guardado,
# OU se ele clicou no menu principal para sair de uma tela de serviço (override ativo)
if radio_val is not None:
    if radio_val != st.session_state.radio_selecionado or st.session_state.menu_override is not None:
        st.session_state.radio_selecionado = radio_val
        st.session_state.menu_override = None
        st.rerun()

# Definimos a variável final 'menu' a ser usada nas telas
menu = st.session_state.menu_override if st.session_state.menu_override else radio_val

# Se o menu mudou, resetamos as sub-telas de navegação das outras páginas para a listagem inicial
if st.session_state.ultimo_menu_acessado != menu:
    st.session_state.sub_tela_mun = "listar"
    st.session_state.sub_tela_bai = "listar"
    st.session_state.sub_tela_upm = "listar"
    st.session_state.sub_tela_ser = "listar"
    st.session_state.ultimo_menu_acessado = menu

# --- EXIBIÇÃO DE SERVIÇOS CADASTRADOS NO SIDEBAR ---
st.sidebar.markdown("<hr style='margin: 15px 0px 15px 0px;'>", unsafe_allow_html=True)
st.sidebar.subheader("🔌 Serviços Ativos")
df_servicos_sidebar = db.listar_dados("servicos")
if not df_servicos_sidebar.empty:
    # Filtra apenas os serviços ativos e que estão marcados para exibição no menu lateral
    df_servicos_sidebar = df_servicos_sidebar[df_servicos_sidebar.get("Status", "Ativo") == "Ativo"]
    if "Exibir_No_Menu" in df_servicos_sidebar.columns:
        df_servicos_sidebar = df_servicos_sidebar[df_servicos_sidebar["Exibir_No_Menu"] == "Sim"]
    
    if not df_servicos_sidebar.empty:
        for idx, row in df_servicos_sidebar.iterrows():
            id_servico = row["ID"]
            nome_servico = row["Nome"]
            is_ativo = (st.session_state.menu_override == f"servico_{id_servico}")
            
            btn_label = f"🎯 {nome_servico} (Ativo)" if is_ativo else f"🔹 {nome_servico}"
            
            if st.sidebar.button(btn_label, key=f"btn_sid_{id_servico}", width="stretch"):
                st.session_state.menu_override = f"servico_{id_servico}"
                st.rerun()
    else:
        st.sidebar.info("Nenhum serviço ativo cadastrado.")
else:
    st.sidebar.info("Nenhum serviço cadastrado.")

st.sidebar.markdown("<hr style='margin: 15px 0px 15px 0px;'>", unsafe_allow_html=True)
st.sidebar.write(f"👤 Logado como: **{user['name']}**")
logout_url = auth_utils.get_logout_url(user.get("id_token", ""))
st.sidebar.link_button("Sair do Sistema (Logout)", logout_url, use_container_width=True)

# Se um serviço estiver ativo via override, vamos injetar CSS para desmarcar visualmente o radio button
if st.session_state.menu_override:
    st.markdown("""
        <style>
            div[data-testid="stSidebarUserContent"] div[role="radiogroup"] label[data-checked="true"] {
                background-color: #f8f9fa !important;
                border-color: #e9ecef !important;
                color: inherit !important;
                font-weight: normal !important;
            }
        </style>
    """, unsafe_allow_html=True)

# =====================================================================
# 1. TELA: MANUTENÇÃO DE MUNICÍPIO (COMPLETA)
# =====================================================================
if menu == "🏙️ Manutenção de Municípios":
    st.title("📍 Manutenção de Município")
    
    if st.session_state.mensagem_sucesso:
        st.success(st.session_state.mensagem_sucesso)
        st.session_state.mensagem_sucesso = None
        
    if st.session_state.sub_tela_mun == "listar":
        if st.button("➕ Cadastrar Novo Município", width="stretch"):
            st.session_state.sub_tela_mun = "formulario"
            st.session_state.modo_form_mun = "cadastro"
            st.session_state.dados_sel_mun = {"ID": None, "Municipio": "", "Estado": "MT - Mato Grosso", "id_municipio_srop": ""}
            st.rerun()
            
        st.write("")
        df_mun = db.listar_dados("municipios")
        
        texto_busca_mun = st.text_input("🔍 Buscar Município por Nome:", key="busca_mun_lista")
        if texto_busca_mun:
            import unicodedata
            def remover_acento_mun(t):
                if not t: return ""
                return "".join(c for c in unicodedata.normalize('NFKD', str(t)) if not unicodedata.combining(c)).lower()
            
            busca_norm = remover_acento_mun(texto_busca_mun)
            mun_norm = df_mun["Municipio"].apply(remover_acento_mun)
            df_mun = df_mun[mun_norm.str.contains(busca_norm, na=False)]
            
        if "pagina_mun" not in st.session_state: st.session_state.pagina_mun = 1
        
        itens_por_pagina = 10
        total_paginas_mun = max(1, (len(df_mun) - 1) // itens_por_pagina + 1)
        if st.session_state.pagina_mun > total_paginas_mun: st.session_state.pagina_mun = total_paginas_mun
        
        start_idx = (st.session_state.pagina_mun - 1) * itens_por_pagina
        end_idx = start_idx + itens_por_pagina
        df_mun_exibicao = df_mun.iloc[start_idx:end_idx]
        
        if df_mun.empty:
            st.info("Nenhum município cadastrado no banco de dados ainda.")
        else:
            st.subheader(f"Municípios Cadastrados (Página {st.session_state.pagina_mun} de {total_paginas_mun})")
            col_id, col_nome, col_est, col_srop, col_acoes = st.columns([0.8, 3.0, 2.0, 2.0, 3.2])
            with col_id: st.write("**ID**")
            with col_nome: st.write("**Município**")
            with col_est: st.write("**Estado**")
            with col_srop: st.write("**ID SROP**")
            with col_acoes: st.write("**Ações Disponíveis**")
            st.markdown("<hr style='margin: 0px 0px 10px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
            
            for idx, row in df_mun_exibicao.iterrows():
                id_atual = row["ID"]
                mun_atual = row["Municipio"]
                est_bruto = str(row["Estado"]).strip()
                mapa_estados_aux = {
                    "AC": "Acre", "AL": "Alagoas", "AP": "Amapá", "AM": "Amazonas",
                    "BA": "Bahia", "CE": "Ceará", "DF": "Distrito Federal", "ES": "Espírito Santo",
                    "GO": "Goiás", "MA": "Maranhão", "MT": "Mato Grosso", "MS": "Mato Grosso do Sul",
                    "MG": "Minas Gerais", "PA": "Pará", "PB": "Paraíba", "PR": "Paraná",
                    "PE": "Pernambuco", "PI": "Piauí", "RJ": "Rio de Janeiro", "RN": "Rio Grande do Norte",
                    "RS": "Rio Grande do Sul", "RO": "Rondônia", "RR": "Roraima", "SC": "Santa Catarina",
                    "SP": "São Paulo", "SE": "Sergipe", "TO": "Tocantins"
                }
                if len(est_bruto) == 2:
                    nome_uf = mapa_estados_aux.get(est_bruto.upper(), "")
                    est_atual = f"{est_bruto.upper()} - {nome_uf}" if nome_uf else est_bruto
                else:
                    est_atual = est_bruto
                    
                srop_atual = row.get("id_municipio_srop", "")
                if pd.isna(srop_atual) or srop_atual is None:
                    srop_atual = ""
                
                c_id, c_nome, c_est, c_srop, c_vis, c_edt, c_exc = st.columns([0.8, 3.0, 2.0, 2.0, 1.0, 1.0, 1.2])
                c_id.write(f"`{id_atual}`")
                c_nome.write(mun_atual)
                c_est.write(est_atual)
                c_srop.write(str(srop_atual) if srop_atual else "-")
                
                if c_vis.button("👁️ Ver", key=f"vis_mun_{id_atual}", width="stretch"):
                    st.session_state.sub_tela_mun = "formulario"; st.session_state.modo_form_mun = "visualizar"
                    st.session_state.dados_sel_mun = {"ID": id_atual, "Municipio": mun_atual, "Estado": est_atual, "id_municipio_srop": srop_atual}; st.rerun()
                    
                if c_edt.button("✏️ Editar", key=f"edt_mun_{id_atual}", width="stretch"):
                    st.session_state.sub_tela_mun = "formulario"; st.session_state.modo_form_mun = "editar"
                    st.session_state.dados_sel_mun = {"ID": id_atual, "Municipio": mun_atual, "Estado": est_atual, "id_municipio_srop": srop_atual}; st.rerun()
                    
                if c_exc.button("🗑️ Excluir", key=f"exc_mun_{id_atual}", width="stretch", type="primary"):
                    st.session_state.sub_tela_mun = "formulario"; st.session_state.modo_form_mun = "excluir"
                    st.session_state.dados_sel_mun = {"ID": id_atual, "Municipio": mun_atual, "Estado": est_atual, "id_municipio_srop": srop_atual}; st.rerun()
            
            st.write("")
            col_ant, col_e1, col_pag, col_e2, col_prox = st.columns([1, 1, 0.8, 1, 1])
            if col_ant.button("⬅️ Página Anterior", disabled=(st.session_state.pagina_mun <= 1), use_container_width=True, key="btn_ant_mun"):
                st.session_state.pagina_mun -= 1
                st.session_state.combo_pag_mun = st.session_state.pagina_mun
                st.rerun()
                
            def set_page_mun():
                st.session_state.pagina_mun = st.session_state.combo_pag_mun
                
            with col_pag:
                st.selectbox(
                    "Pular para a página", 
                    options=range(1, total_paginas_mun + 1), 
                    index=st.session_state.pagina_mun - 1,
                    key="combo_pag_mun",
                    on_change=set_page_mun,
                    label_visibility="collapsed"
                )
                
            if col_prox.button("Próxima Página ➡️", disabled=(st.session_state.pagina_mun >= total_paginas_mun), use_container_width=True, key="btn_prox_mun"):
                st.session_state.pagina_mun += 1
                st.session_state.combo_pag_mun = st.session_state.pagina_mun
                st.rerun()

    elif st.session_state.sub_tela_mun == "formulario":
        modo = st.session_state.modo_form_mun; dados = st.session_state.dados_sel_mun
        campos_bloqueados = True if modo in ["visualizar", "excluir"] else False
        
        if modo == "cadastro": st.subheader("➕ Cadastrar Novo Município")
        elif modo == "visualizar": st.subheader(f"👁️ Visualizando Registro ID: {dados['ID']}")
        elif modo == "editar": st.subheader(f"✏️ Editando Registro ID: {dados['ID']}")
        elif modo == "excluir":
            st.subheader(f"⚠️ Confirmar Exclusão do Registro ID: {dados['ID']}")
            st.error(f"Atenção: Você está prestes a deletar o município '{dados['Municipio']}'. Esta ação não pode ser desfeita.")

        novo_mun = st.text_input("Nome do Município", value=dados["Municipio"], disabled=campos_bloqueados, key="input_mun_dinamico").strip()
        estado_mun = st.selectbox("Estado", ["MT - Mato Grosso"], index=0, disabled=campos_bloqueados, key="select_est_dinamico")
        id_municipio_srop_val = st.text_input("Código ID Município SROP", value=str(dados.get("id_municipio_srop", "")) if dados.get("id_municipio_srop") is not None else "", disabled=campos_bloqueados, key="input_mun_srop_id").strip()
        
        st.write("")
        if modo == "cadastro":
            c_salvar, c_cancelar = st.columns(2)
            if c_salvar.button("💾 Salvar no Banco", key="btn_salvar_mun", width="stretch"):
                if novo_mun:
                    if db.salvar_registro("municipios", {"Municipio": novo_mun, "Estado": estado_mun, "id_municipio_srop": id_municipio_srop_val if id_municipio_srop_val else None}):
                        st.session_state.sub_tela_mun = "listar"; st.session_state.mensagem_sucesso = f"🎉 Município '{novo_mun}' cadastrado com sucesso!"; st.rerun()
                    else: st.error(f"⚠️ Erro: O município '{novo_mun}' já existe!")
                else: st.error("Por favor, digite o nome do município.")
            if c_cancelar.button("❌ Cancelar e Voltar", key="btn_cancelar_cad_mun", width="stretch"):
                st.session_state.sub_tela_mun = "listar"; st.rerun()
                    
        elif modo == "visualizar":
            if st.button("⬅️ Voltar para a Consulta", key="btn_voltar_vis_mun", width="stretch"):
                st.session_state.sub_tela_mun = "listar"; st.rerun()
                
        elif modo == "editar":
            c_atualizar, c_cancelar = st.columns(2)
            if c_atualizar.button("💾 Salvar Alterações", key="btn_update_mun", width="stretch"):
                if novo_mun:
                    conn = db.obter_conexao(); cursor = conn.cursor()
                    cursor.execute("UPDATE municipios SET Municipio = %s, Estado = %s, id_municipio_srop = %s WHERE ID = %s", (novo_mun, estado_mun, id_municipio_srop_val if id_municipio_srop_val else None, dados["ID"]))
                    conn.commit(); conn.close()
                    st.session_state.sub_tela_mun = "listar"; st.session_state.mensagem_sucesso = "✏️ Município alterado com sucesso!"; st.rerun()
                else: st.error("O nome do município não pode ficar em branco.")
            if c_cancelar.button("❌ Cancelar", key="btn_cancelar_edit_mun", width="stretch"):
                st.session_state.sub_tela_mun = "listar"; st.rerun()
                
        elif modo == "excluir":
            c_deletar, c_voltar = st.columns(2)
            if c_deletar.button("🗑️ Sim, Excluir Registro", key="btn_delete_confirm_mun", width="stretch"):
                db.excluir_registro("municipios", dados["ID"])
                st.session_state.sub_tela_mun = "listar"; st.session_state.mensagem_sucesso = f"🗑️ Município '{dados['Municipio']}' removido!"; st.rerun()
            if c_voltar.button("❌ Cancelar e Manter", key="btn_voltar_del_mun", width="stretch"):
                st.session_state.sub_tela_mun = "listar"; st.rerun()

# =====================================================================
# 2. TELA: MANUTENÇÃO DE BAIRRO (COM FILTROS AVANÇADOS)
# =====================================================================
elif menu == "🏘️ Manutenção de Bairros":
    st.title("🏡 Manutenção de Bairro")
    
    if st.session_state.mensagem_sucesso:
        st.success(st.session_state.mensagem_sucesso)
        st.session_state.mensagem_sucesso = None
        
    if st.session_state.sub_tela_bai == "listar":
        if st.button("➕ Cadastrar Novo Bairro", width="stretch"):
            st.session_state.sub_tela_bai = "formulario"
            st.session_state.modo_form_bai = "cadastro"
            st.session_state.dados_sel_bai = {"ID": None, "Bairro": "", "Municipio": ""}
            st.rerun()
            
        st.write("")
        df_bai = db.listar_bairros_com_municipio()
        df_mun_existentes = db.listar_dados("municipios")
        
        if df_bai.empty:
            st.info("Nenhum bairro cadastrado no banco de dados ainda.")
        else:
            # --- SEÇÃO DE FILTROS COM LAYOUT APRIMORADO ---
            with st.container():
                # Proporções otimizadas para os filtros ficarem alinhados lado a lado
                col_filtro_mun, col_filtro_txt = st.columns([2, 3])
                
                with col_filtro_mun:
                    lista_mun_filtro = ["Todos os Municípios"]
                    if not df_mun_existentes.empty:
                        lista_mun_filtro.extend(df_mun_existentes["Municipio"].unique().tolist())
                    
                    if "mem_filtro_bai_mun" not in st.session_state:
                        st.session_state.mem_filtro_bai_mun = "Todos os Municípios"
                    idx_mun = 0
                    if st.session_state.mem_filtro_bai_mun in lista_mun_filtro:
                        idx_mun = lista_mun_filtro.index(st.session_state.mem_filtro_bai_mun)
                        
                    # Reduzimos o rótulo incluindo um emoji para um visual mais limpo
                    municipio_filtrado = st.selectbox("🏙️ Filtrar por Município", lista_mun_filtro, index=idx_mun, key="filtro_mun_bai")
                    st.session_state.mem_filtro_bai_mun = municipio_filtrado
                    
                with col_filtro_txt:
                    if "mem_filtro_bai_txt" not in st.session_state:
                        st.session_state.mem_filtro_bai_txt = ""
                        
                    texto_filtrado = st.text_input("🔍 Digite o nome do Bairro para pesquisar...", value=st.session_state.mem_filtro_bai_txt, key="filtro_txt_bai").strip()
                    st.session_state.mem_filtro_bai_txt = texto_filtrado
            
            st.markdown("<hr style='margin: 10px 0px 20px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
            
            # --- APLICAÇÃO DOS FILTROS NO DATAFRAME ---
            # 1. Filtro por Município (Selectbox)
            if municipio_filtrado != "Todos os Municípios":
                df_bai = df_bai[df_bai["Municipio"] == municipio_filtrado]
                
            # 2. Filtro por Texto (Ignorando Case e Acentuação)
            if texto_filtrado:
                import unicodedata
                
                def remover_acentos_e_case(texto):
                    if not texto:
                        return ""
                    # Remove acentuações e transforma em minúsculo
                    nfkd_form = unicodedata.normalize('NFKD', str(texto))
                    return "".join([c for c in nfkd_form if not unicodedata.combining(c)]).lower()
                
                # Cria uma busca temporária normalizada no Pandas
                bairros_normalizados = df_bai["Bairro"].apply(remover_acentos_e_case)
                termo_busca_normalizado = remover_acentos_e_case(texto_filtrado)
                
                # Filtra as linhas que contêm o termo pesquisado
                df_bai = df_bai[bairros_normalizados.str.contains(termo_busca_normalizado, na=False)]
            
            if "pagina_bai" not in st.session_state: st.session_state.pagina_bai = 1
            
            itens_por_pagina_bai = 10
            total_paginas_bai = max(1, (len(df_bai) - 1) // itens_por_pagina_bai + 1)
            if st.session_state.pagina_bai > total_paginas_bai: st.session_state.pagina_bai = total_paginas_bai
            
            start_idx_bai = (st.session_state.pagina_bai - 1) * itens_por_pagina_bai
            end_idx_bai = start_idx_bai + itens_por_pagina_bai
            df_bai_exibicao = df_bai.iloc[start_idx_bai:end_idx_bai]
            
            st.write("")
            
            if df_bai.empty:
                st.warning("Nenhum bairro encontrado para os filtros selecionados.")
            else:
                st.subheader(f"Bairros Cadastrados (Página {st.session_state.pagina_bai} de {total_paginas_bai})")
                col_id, col_nome, col_mun, col_est, col_acoes = st.columns([0.6, 2.2, 2.2, 1.6, 5.4])
                with col_id: st.write("**ID**")
                with col_nome: st.write("**Bairro**")
                with col_mun: st.write("**Município Vinculado**")
                with col_est: st.write("**Estado**")
                with col_acoes: st.write("**Ações Disponíveis**")
                st.markdown("<hr style='margin: 0px 0px 10px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
                
                for idx, row in df_bai_exibicao.iterrows():
                    id_atual = row["ID"]
                    bai_atual = row["Bairro"]
                    mun_atual = row["Municipio"]
                    
                    # Obtém o estado correspondente a partir do JOIN
                    est_atual = row.get("Estado", "-")
                    
                    c_id, c_nome, c_mun, c_est, c_vis, c_edt, c_alt, c_exc = st.columns([0.6, 2.2, 2.2, 1.6, 1.2, 1.3, 1.3, 1.6])
                    c_id.write(f"`{id_atual}`")
                    c_nome.write(bai_atual)
                    c_mun.text(mun_atual)
                    c_est.write(est_atual)
                    
                    if c_vis.button("👁️ Ver", key=f"vis_bai_{id_atual}", width="stretch"):
                        st.session_state.sub_tela_bai = "formulario"; st.session_state.modo_form_bai = "visualizar"
                        st.session_state.dados_sel_bai = {"ID": id_atual, "Bairro": bai_atual, "Municipio": mun_atual}; st.rerun()
                    if c_edt.button("✏️ Editar", key=f"edt_bai_{id_atual}", width="stretch"):
                        st.session_state.sub_tela_bai = "formulario"; st.session_state.modo_form_bai = "editar"
                        st.session_state.dados_sel_bai = {"ID": id_atual, "Bairro": bai_atual, "Municipio": mun_atual}; st.rerun()
                    if c_alt.button("🏷️ Nomes", key=f"alt_bai_{id_atual}", width="stretch"):
                        st.session_state.sub_tela_bai = "alternativos"
                        st.session_state.dados_sel_bai = {"ID": id_atual, "Bairro": bai_atual, "Municipio": mun_atual}; st.rerun()
                    if c_exc.button("🗑️ Excluir", key=f"exc_bai_{id_atual}", width="stretch", type="primary"):
                        st.session_state.sub_tela_bai = "formulario"; st.session_state.modo_form_bai = "excluir"
                        st.session_state.dados_sel_bai = {"ID": id_atual, "Bairro": bai_atual, "Municipio": mun_atual}; st.rerun()
                
                st.write("")
                col_ant, col_e1, col_pag, col_e2, col_prox = st.columns([1, 1, 0.8, 1, 1])
                if col_ant.button("⬅️ Página Anterior", disabled=(st.session_state.pagina_bai <= 1), use_container_width=True, key="btn_ant_bai"):
                    st.session_state.pagina_bai -= 1
                    st.session_state.combo_pag_bai = st.session_state.pagina_bai
                    st.rerun()
                    
                def set_page_bai():
                    st.session_state.pagina_bai = st.session_state.combo_pag_bai
                    
                with col_pag:
                    st.selectbox(
                        "Pular para a página", 
                        options=range(1, total_paginas_bai + 1), 
                        index=st.session_state.pagina_bai - 1,
                        key="combo_pag_bai",
                        on_change=set_page_bai,
                        label_visibility="collapsed"
                    )
                    
                if col_prox.button("Próxima Página ➡️", disabled=(st.session_state.pagina_bai >= total_paginas_bai), use_container_width=True, key="btn_prox_bai"):
                    st.session_state.pagina_bai += 1
                    st.session_state.combo_pag_bai = st.session_state.pagina_bai
                    st.rerun()

    elif st.session_state.sub_tela_bai == "formulario":
        modo = st.session_state.modo_form_bai; dados = st.session_state.dados_sel_bai
        campos_bloqueados = True if modo in ["visualizar", "excluir"] else False
        
        if modo == "cadastro": st.subheader("➕ Cadastrar Novo Bairro")
        elif modo == "visualizar": st.subheader(f"👁️ Visualizando Bairro ID: {dados['ID']}")
        elif modo == "editar": st.subheader(f"✏️ Editando Bairro ID: {dados['ID']}")
        elif modo == "excluir":
            st.subheader(f"⚠️ Confirmar Exclusão do Bairro ID: {dados['ID']}")
            st.error(f"Atenção: Você está prestes a remover o bairro '{dados['Bairro']}' de '{dados['Municipio']}'. Esta ação não pode ser desfeita.")

        novo_bairro = st.text_input("Nome do Bairro", value=dados["Bairro"], disabled=campos_bloqueados, key="input_bai_dinamico").strip()
        
        df_mun_existentes = db.listar_dados("municipios")
        lista_formatada = []
        
        if not df_mun_existentes.empty:
            df_mun_existentes["UF"] = df_mun_existentes["Estado"].str[:2]
            df_mun_existentes["Exibicao"] = df_mun_existentes["Municipio"] + " / " + df_mun_existentes["UF"]
            lista_formatada = df_mun_existentes["Exibicao"].tolist()
        
        idx_padrao = 0
        if dados["Municipio"] and not df_mun_existentes.empty:
            filtro = df_mun_existentes[df_mun_existentes["Municipio"] == dados["Municipio"]]
            if not filtro.empty:
                str_exibicao_salva = filtro.iloc[0]["Exibicao"]
                if str_exibicao_salva in lista_formatada:
                    idx_padrao = lista_formatada.index(str_exibicao_salva)

        municipio_selecionado_combo = st.selectbox(
            "Vincular ao Município", 
            lista_formatada if lista_formatada else ["Cadastre um município primeiro antes de continuar"], 
            index=idx_padrao, 
            disabled=campos_bloqueados, 
            key="sel_mun_bai_form"
        )
        
        st.write("")
        if modo == "cadastro":
            c_s, c_c = st.columns(2)
            if c_s.button("💾 Salvar Bairro", key="btn_salvar_bairro_db", width="stretch"):
                if novo_bairro and lista_formatada:
                    # Resolve o ID do município selecionado
                    municipio_puro = municipio_selecionado_combo.split(" / ")[0].strip()
                    filtro_id = df_mun_existentes[df_mun_existentes["Municipio"] == municipio_puro]
                    municipio_id = int(filtro_id.iloc[0]["ID"]) if not filtro_id.empty else None
                    
                    if municipio_id:
                        if db.salvar_registro("bairros", {"Bairro": novo_bairro, "Municipio_ID": municipio_id}):
                            st.session_state.sub_tela_bai = "listar"
                            st.session_state.mensagem_sucesso = f"🎉 Bairro '{novo_bairro}' cadastrado com sucesso!"
                            st.rerun()
                        else: 
                            st.error("⚠️ Erro: Este bairro já está cadastrado para o município selecionado!")
                    else:
                        st.error("⚠️ Município inválido. Selecione um município válido.")
                else: 
                    st.error("Por favor, preencha o nome do bairro.")
            if c_c.button("❌ Cancelar e Voltar", key="btn_cancelar_cad_bai", width="stretch"): 
                st.session_state.sub_tela_bai = "listar"; st.rerun()
                
        elif modo == "visualizar":
            if st.button("⬅️ Voltar para a Consulta", key="btn_voltar_vis_bai", width="stretch"): 
                st.session_state.sub_tela_bai = "listar"; st.rerun()
                
        elif modo == "editar":
            c_up, c_cc = st.columns(2)
            if c_up.button("💾 Salvar Alterações", key="btn_update_bairro_db", width="stretch"):
                if novo_bairro:
                    # Resolve o ID do município selecionado
                    municipio_puro = municipio_selecionado_combo.split(" / ")[0].strip()
                    filtro_id = df_mun_existentes[df_mun_existentes["Municipio"] == municipio_puro]
                    municipio_id = int(filtro_id.iloc[0]["ID"]) if not filtro_id.empty else None
                    
                    if municipio_id:
                        conn = db.obter_conexao(); cursor = conn.cursor()
                        cursor.execute("UPDATE bairros SET Bairro = %s, Municipio_ID = %s WHERE ID = %s", (novo_bairro, municipio_id, dados["ID"]))
                        conn.commit(); conn.close()
                        st.session_state.sub_tela_bai = "listar"
                        st.session_state.mensagem_sucesso = "✏️ Bairro atualizado com sucesso!"
                        st.rerun()
                    else:
                        st.error("⚠️ Município inválido. Selecione um município válido.")
                else:
                    st.error("O nome do bairro não pode ficar em branco.")
            if c_cc.button("❌ Cancelar", key="btn_cancelar_edit_bai", width="stretch"): 
                st.session_state.sub_tela_bai = "listar"; st.rerun()
                
        elif modo == "excluir":
            c_del, c_v = st.columns(2)
            if c_del.button("🗑️ Sim, Excluir Registro", key="btn_delete_confirm_bai", width="stretch"):
                db.excluir_registro("bairros", dados["ID"])
                st.session_state.sub_tela_bai = "listar"
                st.session_state.mensagem_sucesso = f"🗑️ Bairro '{dados['Bairro']}' removido com sucesso!"
                st.rerun()
            if c_v.button("❌ Cancelar e Manter", key="btn_voltar_del_bai", width="stretch"):
                st.session_state.sub_tela_bai = "listar"
                st.rerun()

    elif st.session_state.sub_tela_bai == "alternativos":
        dados = st.session_state.dados_sel_bai
        st.subheader(f"🏷️ Nomes Alternativos para Bairro: {dados['Bairro']}")
        
        st.write("#### Bairro Selecionado")
        col_b_nome, col_b_mun = st.columns(2)
        col_b_nome.text_input("Bairro", value=dados["Bairro"], disabled=True, key="bai_alt_view_nome")
        col_b_mun.text_input("Município", value=dados["Municipio"], disabled=True, key="bai_alt_view_mun")
        
        st.markdown("<hr style='margin: 10px 0px 20px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
        
        # Formulário para cadastrar nova equivalência
        st.write("#### ➕ Cadastrar Novo Nome Alternativo / Equivalente")
        col_input_alt, col_btn_add = st.columns([3, 1])
        
        with col_input_alt:
            novo_alt = st.text_input("Ex: JD SAO SIMAU, SAO SIMAO, S SIMAO", key="input_novo_alt_bairro", label_visibility="collapsed").strip()
            
        with col_btn_add:
            if st.button("➕ Adicionar", key="btn_add_alt_bairro", width="stretch"):
                if novo_alt:
                    # Tenta salvar no banco
                    if db.salvar_nome_alternativo(dados["ID"], novo_alt):
                        st.success(f"🎉 Nome alternativo '{novo_alt}' cadastrado com sucesso!")
                        st.rerun()
                    else:
                        st.error(f"⚠️ Erro: O nome alternativo '{novo_alt}' já está cadastrado para este bairro.")
                else:
                    st.error("Por favor, digite um nome alternativo.")
                    
        st.write("")
        st.write("#### 📋 Nomes Alternativos Cadastrados")
        
        df_alt = db.listar_nomes_alternativos(dados["ID"])
        
        if df_alt.empty:
            st.info("Nenhum nome alternativo cadastrado para este bairro ainda.")
        else:
            col_h_nome, col_h_acao = st.columns([3, 1])
            with col_h_nome: st.write("**Nome Alternativo / Equivalência**")
            with col_h_acao: st.write("**Ação**")
            st.markdown("<hr style='margin: 0px 0px 10px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
            
            for idx, row in df_alt.iterrows():
                alt_id = row["ID"]
                alt_nome = row["Nome_Alternativo"]
                
                c_nome, c_acao = st.columns([3, 1])
                c_nome.write(alt_nome)
                if c_acao.button("🗑️ Excluir", key=f"btn_del_alt_{alt_id}", width="stretch", type="primary"):
                    db.excluir_nome_alternativo(alt_id)
                    st.rerun()
                    
        st.markdown("<hr style='margin: 20px 0px 20px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
        
        if st.button("⬅️ Voltar para a Consulta de Bairros", key="btn_voltar_alt_bai_list", width="stretch"):
            st.session_state.sub_tela_bai = "listar"
            st.rerun()

# =====================================================================
# 3. TELA: MANUTENÇÃO DE UPMS (COMPLETA)
# =====================================================================
elif menu == "🚔 Manutenção de UPMs":
    st.title("🚔 Manutenção de UPMs")
    
    if st.session_state.mensagem_sucesso:
        st.success(st.session_state.mensagem_sucesso)
        st.session_state.mensagem_sucesso = None
        
    if st.session_state.sub_tela_upm == "listar":
        if st.button("➕ Cadastrar Nova UPM", width="stretch"):
            st.session_state.sub_tela_upm = "formulario"; st.session_state.modo_form_upm = "cadastro"
            st.session_state.dados_sel_upm = {"ID": None, "UPM": "", "Descricao": "", "Bairro": "", "Municipio": "", "Estado": ""}; st.rerun()
            
        st.write("")
        df_upm = db.listar_dados("upms")
        if not df_upm.empty:
            import re
            def extrair_numero_para_ordenacao(series):
                def converter_valor(val):
                    val_str = str(val).strip()
                    match = re.search(r'\d+', val_str)
                    num = int(match.group()) if match else 999999
                    return (num, val_str.lower())
                return series.apply(converter_valor)
            df_upm = df_upm.sort_values(by="UPM", key=extrair_numero_para_ordenacao)
        
        if "pagina_upm" not in st.session_state: st.session_state.pagina_upm = 1
        
        itens_por_pagina_upm = 10
        total_paginas_upm = max(1, (len(df_upm) - 1) // itens_por_pagina_upm + 1)
        if st.session_state.pagina_upm > total_paginas_upm: st.session_state.pagina_upm = total_paginas_upm
        
        start_idx_upm = (st.session_state.pagina_upm - 1) * itens_por_pagina_upm
        end_idx_upm = start_idx_upm + itens_por_pagina_upm
        df_upm_exibicao = df_upm.iloc[start_idx_upm:end_idx_upm]
        
        if df_upm.empty:
            st.info("Nenhuma UPM cadastrada no banco de dados ainda.")
        else:
            st.subheader(f"UPMs Cadastradas (Página {st.session_state.pagina_upm} de {total_paginas_upm})")
            col_id, col_upm, col_desc, col_vis, col_edt, col_vinc, col_exc = st.columns([0.8, 2.2, 3.4, 1.1, 1.1, 1.3, 1.1])
            with col_id: st.write("**ID**")
            with col_upm: st.write("**UPM**")
            with col_desc: st.write("**Descrição**")
            with col_vis: st.write("**Ver**")
            with col_edt: st.write("**Editar**")
            with col_vinc: st.write("**Vincular**")
            with col_exc: st.write("**Excluir**")
            st.markdown("<hr style='margin: 0px 0px 10px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
            
            for idx, row in df_upm_exibicao.iterrows():
                id_atual = row["ID"]
                upm_atual = row["UPM"]
                desc_atual = row.get("Descricao", "")
                
                c_id, c_upm, c_desc, c_vis, c_edt, c_vinc, c_exc = st.columns([0.8, 2.2, 3.4, 1.1, 1.1, 1.3, 1.1])
                c_id.write(f"`{id_atual}`")
                c_upm.write(upm_atual)
                c_desc.write(desc_atual)
                
                if c_vis.button("👁️ Ver", key=f"vis_upm_{id_atual}", width="stretch"):
                    st.session_state.sub_tela_upm = "formulario"; st.session_state.modo_form_upm = "visualizar"
                    st.session_state.dados_sel_upm = {"ID": id_atual, "UPM": upm_atual, "Descricao": desc_atual}; st.rerun()
                if c_edt.button("✏️ Editar", key=f"edt_upm_{id_atual}", width="stretch"):
                    st.session_state.sub_tela_upm = "formulario"; st.session_state.modo_form_upm = "editar"
                    st.session_state.dados_sel_upm = {"ID": id_atual, "UPM": upm_atual, "Descricao": desc_atual}; st.rerun()
                if c_vinc.button("🔗 Vincular", key=f"vinc_upm_{id_atual}", width="stretch"):
                    st.session_state.sub_tela_upm = "vincular"
                    st.session_state.dados_sel_upm = {"ID": id_atual, "UPM": upm_atual, "Descricao": desc_atual}; st.rerun()
                if c_exc.button("🗑️ Excluir", key=f"exc_upm_{id_atual}", width="stretch", type="primary"):
                    st.session_state.sub_tela_upm = "formulario"; st.session_state.modo_form_upm = "excluir"
                    st.session_state.dados_sel_upm = {"ID": id_atual, "UPM": upm_atual, "Descricao": desc_atual}; st.rerun()
            
            st.write("")
            col_ant, col_e1, col_pag, col_e2, col_prox = st.columns([1, 1, 0.8, 1, 1])
            if col_ant.button("⬅️ Página Anterior", disabled=(st.session_state.pagina_upm <= 1), use_container_width=True, key="btn_ant_upm"):
                st.session_state.pagina_upm -= 1
                st.session_state.combo_pag_upm = st.session_state.pagina_upm
                st.rerun()
                
            def set_page_upm():
                st.session_state.pagina_upm = st.session_state.combo_pag_upm
                
            with col_pag:
                st.selectbox(
                    "Pular para a página", 
                    options=range(1, total_paginas_upm + 1), 
                    index=st.session_state.pagina_upm - 1,
                    key="combo_pag_upm",
                    on_change=set_page_upm,
                    label_visibility="collapsed"
                )
                
            if col_prox.button("Próxima Página ➡️", disabled=(st.session_state.pagina_upm >= total_paginas_upm), use_container_width=True, key="btn_prox_upm"):
                st.session_state.pagina_upm += 1
                st.session_state.combo_pag_upm = st.session_state.pagina_upm
                st.rerun()

    elif st.session_state.sub_tela_upm == "formulario":
        modo = st.session_state.modo_form_upm; dados = st.session_state.dados_sel_upm
        campos_bloqueados = True if modo in ["visualizar", "excluir"] else False
        
        if modo == "cadastro": st.subheader("➕ Cadastrar Nova UPM")
        elif modo == "visualizar": st.subheader(f"👁️ Visualizando UPM ID: {dados['ID']}")
        elif modo == "editar": st.subheader(f"✏️ Editando UPM ID: {dados['ID']}")
        elif modo == "excluir":
            st.subheader(f"⚠️ Confirmar Exclusão da UPM ID: {dados['ID']}")
            st.error(f"Atenção: Você está prestes a deletar a UPM '{dados['UPM']}'. Esta ação não pode ser desfeita.")

        nome_upm = st.text_input("Nome/Identificador da UPM", value=dados["UPM"], disabled=campos_bloqueados, key="upm_nome_in").strip()
        descricao_upm = st.text_area("Descrição da UPM", value=dados.get("Descricao", ""), disabled=campos_bloqueados, key="upm_desc_in", height=120).strip()
        
        st.write("")
        if modo == "cadastro":
            c_s, c_c = st.columns(2)
            if c_s.button("💾 Salvar no Banco", key="btn_salvar_upm", width="stretch"):
                if nome_upm:
                    if db.salvar_registro("upms", {"UPM": nome_upm, "Descricao": descricao_upm}):
                        st.session_state.sub_tela_upm = "listar"; st.session_state.mensagem_sucesso = f"🎉 UPM '{nome_upm}' cadastrada com sucesso!"; st.rerun()
                    else: st.error(f"⚠️ Erro: A UPM '{nome_upm}' já existe!")
                else: st.error("Por favor, preencha o nome/identificador da UPM.")
            if c_c.button("❌ Cancelar e Voltar", key="btn_cancelar_cad_upm", width="stretch"):
                st.session_state.sub_tela_upm = "listar"; st.rerun()
                
        elif modo == "visualizar":
            if st.button("⬅️ Voltar para a Consulta", key="btn_voltar_vis_upm", width="stretch"):
                st.session_state.sub_tela_upm = "listar"; st.rerun()
                
        elif modo == "editar":
            c_s, c_c = st.columns(2)
            if c_s.button("💾 Salvar Alterações", key="btn_update_upm", width="stretch"):
                if nome_upm:
                    conn = db.obter_conexao(); cursor = conn.cursor()
                    cursor.execute("UPDATE upms SET UPM = %s, Descricao = %s WHERE ID = %s", (nome_upm, descricao_upm, dados["ID"]))
                    conn.commit(); conn.close()
                    st.session_state.sub_tela_upm = "listar"; st.session_state.mensagem_sucesso = "✏️ UPM alterada com sucesso!"; st.rerun()
                else: st.error("O nome da UPM não pode ficar em branco.")
            if c_c.button("❌ Cancelar", key="btn_cancelar_edit_upm", width="stretch"):
                st.session_state.sub_tela_upm = "listar"; st.rerun()
                
        elif modo == "excluir":
            c_s, c_c = st.columns(2)
            if c_s.button("🗑️ Sim, Excluir Registro", key="btn_delete_confirm_upm", width="stretch"):
                # Limpa vínculos em upm_bairros primeiro
                conn = db.obter_conexao(); cursor = conn.cursor()
                cursor.execute("DELETE FROM upm_bairros WHERE UPM_ID = %s", (dados["ID"],))
                conn.commit(); conn.close()
                # Exclui a UPM
                db.excluir_registro("upms", dados["ID"])
                st.session_state.sub_tela_upm = "listar"; st.session_state.mensagem_sucesso = f"🗑️ UPM '{dados['UPM']}' removida!"; st.rerun()
            if c_c.button("❌ Cancelar e Manter", key="btn_voltar_del_upm", width="stretch"):
                st.session_state.sub_tela_upm = "listar"; st.rerun()

    elif st.session_state.sub_tela_upm == "vincular":
        dados = st.session_state.dados_sel_upm
        st.subheader(f"🔗 Vincular Bairros à UPM: {dados['UPM']}")
        
        # Exibe os detalhes da UPM selecionada como somente leitura
        st.write("#### UPM Selecionada")
        col_upm_v, col_desc_v = st.columns([1, 2])
        col_upm_v.text_input("UPM", value=dados["UPM"], disabled=True, key="upm_vinc_view_nome")
        col_desc_v.text_area("Descrição", value=dados.get("Descricao", ""), disabled=True, key="upm_vinc_view_desc", height=68)
        
        df_b = db.listar_bairros_com_municipio(); df_m = db.listar_dados("municipios")
        
        if df_b.empty:
            st.warning("Cadastre bairros antes de criar vínculos.")
            if st.button("⬅️ Voltar", width="stretch", key="upm_vinc_back_empty"):
                st.session_state.sub_tela_upm = "listar"; st.rerun()
        else:
            # Bairros já vem com Municipio e Estado via JOIN — sem necessidade de merge
            df_bairros_completo = df_b.copy()
            df_bairros_completo["Exibicao"] = df_bairros_completo["Bairro"] + " (" + df_bairros_completo["Municipio"] + ")"
            
            # Inicializa a lista de selecionados no session_state para persistência
            if "selected_bairro_ids" not in st.session_state:
                bairros_vinculados = db.listar_bairros_vinculados(dados["ID"])
                st.session_state.selected_bairro_ids = set(bairros_vinculados["BairroID"].tolist())
                st.session_state.vinc_default_mun = "Todos os Municípios"
                
            st.markdown("<hr style='margin: 10px 0px 20px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
            
            # Estrutura de colunas para layout profissional de Filtros e Resumo
            col_filtros, col_resumo = st.columns([3, 1])
            
            with col_filtros:
                st.write("#### 🔍 Filtros de Busca")
                col_filtro_mun, col_filtro_txt = st.columns([1, 1])
                
                with col_filtro_mun:
                    lista_municipios = ["Todos os Municípios"]
                    if not df_m.empty:
                        lista_municipios.extend(df_m["Municipio"].unique().tolist())
                    
                    idx_padrao_mun = 0
                    if "vinc_default_mun" in st.session_state and st.session_state.vinc_default_mun in lista_municipios:
                        idx_padrao_mun = lista_municipios.index(st.session_state.vinc_default_mun)
                        
                    mun_selecionado = st.selectbox(
                        "🏙️ Filtrar por Município", 
                        lista_municipios, 
                        index=idx_padrao_mun, 
                        key="vinc_mun_filtro"
                    )
                    
                with col_filtro_txt:
                    busca_txt = st.text_input("🔍 Buscar por nome do Bairro", key="vinc_busca_txt").strip()
            
            with col_resumo:
                st.write("#### 📍 Resumo")
                st.info(f"Selecionados: **{len(st.session_state.selected_bairro_ids)}** bairros")
                if st.button("🗑️ Limpar Seleções", width="stretch", key="clear_all_vinc"):
                    st.session_state.selected_bairro_ids.clear()
                    st.rerun()

            # Reseta a paginação se os filtros mudarem
            if "ultimo_mun_filtro" not in st.session_state:
                st.session_state.ultimo_mun_filtro = mun_selecionado
            if "ultima_busca_txt" not in st.session_state:
                st.session_state.ultima_busca_txt = busca_txt
                
            if st.session_state.ultimo_mun_filtro != mun_selecionado or st.session_state.ultima_busca_txt != busca_txt:
                st.session_state.pagina_vinc = 1
                st.session_state.ultimo_mun_filtro = mun_selecionado
                st.session_state.ultima_busca_txt = busca_txt

            # Ordena bairros colocando os atualmente selecionados no topo em tempo real
            df_bairros_completo["Vinculado"] = df_bairros_completo["ID"].apply(
                lambda x: x in st.session_state.selected_bairro_ids
            )
            df_ordenado = df_bairros_completo.sort_values(
                by=["Vinculado", "Municipio", "Bairro"],
                ascending=[False, True, True]
            )

            # Aplica os filtros ao DataFrame ordenado
            df_filtrado = df_ordenado.copy()
            if mun_selecionado != "Todos os Municípios":
                df_filtrado = df_filtrado[df_filtrado["Municipio"] == mun_selecionado]
            
            if busca_txt:
                import unicodedata
                def normalizar(t):
                    if not t: return ""
                    nfkd = unicodedata.normalize('NFKD', str(t))
                    return "".join([c for c in nfkd if not unicodedata.combining(c)]).lower()
                
                bairros_norm = df_filtrado["Bairro"].apply(normalizar)
                busca_norm = normalizar(busca_txt)
                df_filtrado = df_filtrado[bairros_norm.str.contains(busca_norm, na=False)]
                
            st.write("")
            st.write("#### 🏡 Lista de Bairros para Vínculo")
            
            if df_filtrado.empty:
                st.warning("Nenhum bairro encontrado com os filtros aplicados.")
            else:
                # Paginação
                itens_por_pagina = 15
                total_itens = len(df_filtrado)
                total_paginas = max((total_itens - 1) // itens_por_pagina + 1, 1)
                
                if "pagina_vinc" not in st.session_state:
                    st.session_state.pagina_vinc = 1
                if st.session_state.pagina_vinc > total_paginas:
                    st.session_state.pagina_vinc = total_paginas
                
                inicio = (st.session_state.pagina_vinc - 1) * itens_por_pagina
                fim = inicio + itens_por_pagina
                df_pagina = df_filtrado.iloc[inicio:fim]

                # Cabeçalhos da tabela interativa
                col_h_status, col_h_bairro, col_h_mun, col_h_acao = st.columns([1.5, 3.5, 3.5, 2.5])
                with col_h_status: st.write("**Status**")
                with col_h_bairro: st.write("**Bairro**")
                with col_h_mun: st.write("**Município**")
                with col_h_acao: st.write("**Ação**")
                st.markdown("<hr style='margin: 0px 0px 10px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)

                # Renderiza cada linha da tabela interativa
                for idx, row in df_pagina.iterrows():
                    b_id = int(row["ID"])
                    b_nome = row["Bairro"]
                    b_mun = row["Municipio"]
                    is_sel = b_id in st.session_state.selected_bairro_ids
                    
                    c_status, c_bairro, c_mun, c_acao = st.columns([1.5, 3.5, 3.5, 2.5])
                    
                    c_bairro.write(b_nome)
                    c_mun.write(b_mun)
                    
                    if is_sel:
                        c_status.write("🔗 **Vinculado**")
                        if c_acao.button("❌ Desvincular", key=f"btn_desv_{b_id}", width="stretch"):
                            st.session_state.selected_bairro_ids.discard(b_id)
                            st.rerun()
                    else:
                        c_status.write("⚪ Não Vinculado")
                        if c_acao.button("➕ Vincular", key=f"btn_vinc_{b_id}", width="stretch"):
                            st.session_state.selected_bairro_ids.add(b_id)
                            st.rerun()
                
                # Controles de paginação
                st.write("")
                col_pag_prev, col_pag_info, col_pag_next = st.columns([1, 2, 1])
                if col_pag_prev.button("⬅️ Anterior", disabled=(st.session_state.pagina_vinc == 1), width="stretch", key="pag_vinc_prev"):
                    st.session_state.pagina_vinc -= 1
                    st.rerun()
                
                col_pag_info.markdown(f"<p style='text-align: center; margin-top: 6px;'>Página <b>{st.session_state.pagina_vinc}</b> de <b>{total_paginas}</b> (Total: {total_itens} bairros)</p>", unsafe_allow_html=True)
                
                if col_pag_next.button("Próxima ➡️", disabled=(st.session_state.pagina_vinc == total_paginas), width="stretch", key="pag_vinc_next"):
                    st.session_state.pagina_vinc += 1
                    st.rerun()
            
            st.markdown("<hr style='margin: 20px 0px 10px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
            
            c_salvar, c_cancelar = st.columns(2)
            if c_salvar.button("💾 Salvar Vínculos", width="stretch", key="upm_vinc_save"):
                db.atualizar_vinculo_bairros(dados["ID"], list(st.session_state.selected_bairro_ids))
                # Limpa estados temporários
                for k in ["selected_bairro_ids", "pagina_vinc", "ultimo_mun_filtro", "ultima_busca_txt", "vinc_mun_filtro", "vinc_default_mun"]:
                    if k in st.session_state:
                        del st.session_state[k]
                st.session_state.sub_tela_upm = "listar"
                st.session_state.mensagem_sucesso = f"🔗 Bairros vinculados à UPM '{dados['UPM']}' com sucesso!"
                st.rerun()
                
            if c_cancelar.button("❌ Cancelar e Voltar", width="stretch", key="upm_vinc_cancel"):
                # Limpa estados temporários
                for k in ["selected_bairro_ids", "pagina_vinc", "ultimo_mun_filtro", "ultima_busca_txt", "vinc_mun_filtro", "vinc_default_mun"]:
                    if k in st.session_state:
                        del st.session_state[k]
                st.session_state.sub_tela_upm = "listar"; st.rerun()

# =====================================================================
# 3.5. TELA: MANUTENÇÃO DE SERVIÇOS (COMPLETA)
# =====================================================================
elif menu == "🌐 Manutenção de Serviços":
    st.title("🌐 Manutenção de Serviços")
    
    if st.session_state.mensagem_sucesso:
        st.success(st.session_state.mensagem_sucesso)
        st.session_state.mensagem_sucesso = None
        
    if st.session_state.sub_tela_ser == "listar":
        col_btn_cad, col_btn_mon = st.columns(2)
        if col_btn_cad.button("➕ Cadastrar Novo Serviço", width="stretch"):
            st.session_state.sub_tela_ser = "formulario"
            st.session_state.modo_form_ser = "cadastro"
            st.session_state.dados_sel_ser = {
                "ID": None, "Nome": "", "UrlLogin": "", "UrlConsulta": "", "UrlPdf": "", "Login": "", "Senha": "",
                "DuplaAutenticacao": "Não", "Tipo": "SROP", "Status": "Ativo", "Layout_ID": 1, "Exibir_No_Menu": "Sim"
            }
            st.rerun()
            
        if col_btn_mon.button("🖥️ Monitoramento de Sessões", width="stretch"):
            st.session_state.sub_tela_ser = "monitoramento"
            st.rerun()
            
        st.write("")
        df_ser = db.listar_dados("servicos")
        
        if "pagina_ser" not in st.session_state: st.session_state.pagina_ser = 1
        
        itens_por_pagina_ser = 10
        total_paginas_ser = max(1, (len(df_ser) - 1) // itens_por_pagina_ser + 1)
        if st.session_state.pagina_ser > total_paginas_ser: st.session_state.pagina_ser = total_paginas_ser
        
        start_idx_ser = (st.session_state.pagina_ser - 1) * itens_por_pagina_ser
        end_idx_ser = start_idx_ser + itens_por_pagina_ser
        df_ser_exibicao = df_ser.iloc[start_idx_ser:end_idx_ser]
        
        if df_ser.empty:
            st.info("Nenhum serviço cadastrado no banco de dados ainda.")
        else:
            st.subheader(f"Serviços Cadastrados (Página {st.session_state.pagina_ser} de {total_paginas_ser})")
            col_id, col_nome, col_status, col_acoes = st.columns([0.8, 3, 2, 4.2])
            with col_id: st.write("**ID**")
            with col_nome: st.write("**Nome do Serviço**")
            with col_status: st.write("**Situação**")
            with col_acoes: st.write("**Ações Disponíveis**")
            st.markdown("<hr style='margin: 0px 0px 10px 0px; border-color: #f0f2f6;'>", unsafe_allow_html=True)
            
            for idx, row in df_ser_exibicao.iterrows():
                id_atual = row["ID"]
                nome_atual = row["Nome"]
                login_atual = row["Login"]
                url_login = row["UrlLogin"]
                url_consulta = row["UrlConsulta"]
                url_pdf = row["UrlPdf"]
                senha_cripto = row["Senha"]
                dupla_autenticacao = row.get("DuplaAutenticacao", "Não")
                tipo_servico = row.get("Tipo", "SROP")
                status_servico = row.get("Status", "Ativo")
                tempo_expiracao = row.get("Tempo_Expiracao_Horas", 4)
                layout_id_db = row.get("Layout_ID", 1)
                exibir_no_menu = row.get("Exibir_No_Menu", "Sim")
                
                c_id, c_nome, c_status, c_vis, c_edt, c_exc = st.columns([0.8, 3, 2, 1.2, 1.2, 1.8])
                c_id.write(f"`{id_atual}`")
                c_nome.write(nome_atual)
                
                # Renderiza status com um badge de cor correspondente
                if status_servico == "Ativo":
                    c_status.markdown("🟢 **Ativo**")
                else:
                    c_status.markdown("🔴 **Inativo**")
                
                if c_vis.button("👁️ Ver", key=f"vis_ser_{id_atual}", width="stretch"):
                    st.session_state.sub_tela_ser = "formulario"
                    st.session_state.modo_form_ser = "visualizar"
                    st.session_state.dados_sel_ser = {
                        "ID": id_atual, "Nome": nome_atual, "UrlLogin": url_login,
                        "UrlConsulta": url_consulta, "UrlPdf": url_pdf, "Login": login_atual,
                        "Senha": db.descriptografar_senha(senha_cripto),
                        "DuplaAutenticacao": dupla_autenticacao,
                        "Tipo": tipo_servico, "Status": status_servico,
                        "Tempo_Expiracao_Horas": tempo_expiracao,
                        "Layout_ID": layout_id_db,
                        "Exibir_No_Menu": exibir_no_menu
                    }
                    st.rerun()
                    
                if c_edt.button("✏️ Editar", key=f"edt_ser_{id_atual}", width="stretch"):
                    st.session_state.sub_tela_ser = "formulario"
                    st.session_state.modo_form_ser = "editar"
                    st.session_state.dados_sel_ser = {
                        "ID": id_atual, "Nome": nome_atual, "UrlLogin": url_login,
                        "UrlConsulta": url_consulta, "UrlPdf": url_pdf, "Login": login_atual,
                        "Senha": db.descriptografar_senha(senha_cripto),
                        "DuplaAutenticacao": dupla_autenticacao,
                        "Tipo": tipo_servico, "Status": status_servico,
                        "Tempo_Expiracao_Horas": tempo_expiracao,
                        "Layout_ID": layout_id_db,
                        "Exibir_No_Menu": exibir_no_menu
                    }
                    st.rerun()
                    
                if c_exc.button("🗑️ Excluir", key=f"exc_ser_{id_atual}", width="stretch", type="primary"):
                    st.session_state.sub_tela_ser = "formulario"
                    st.session_state.modo_form_ser = "excluir"
                    st.session_state.dados_sel_ser = {
                        "ID": id_atual, "Nome": nome_atual, "UrlLogin": url_login,
                        "UrlConsulta": url_consulta, "UrlPdf": url_pdf, "Login": login_atual,
                        "Senha": db.descriptografar_senha(senha_cripto),
                        "DuplaAutenticacao": dupla_autenticacao,
                        "Tipo": tipo_servico, "Status": status_servico,
                        "Tempo_Expiracao_Horas": tempo_expiracao,
                        "Layout_ID": layout_id_db,
                        "Exibir_No_Menu": exibir_no_menu
                    }
                    st.rerun()
                    
            st.write("")
            col_ant, col_e1, col_pag, col_e2, col_prox = st.columns([1, 1, 0.8, 1, 1])
            if col_ant.button("⬅️ Página Anterior", disabled=(st.session_state.pagina_ser <= 1), use_container_width=True, key="btn_ant_ser"):
                st.session_state.pagina_ser -= 1; st.rerun()
                
            def set_page_ser():
                st.session_state.pagina_ser = st.session_state.combo_pag_ser
                
            with col_pag:
                st.selectbox(
                    "Pular para a página", 
                    options=range(1, total_paginas_ser + 1), 
                    index=st.session_state.pagina_ser - 1,
                    key="combo_pag_ser",
                    on_change=set_page_ser,
                    label_visibility="collapsed"
                )
                
            if col_prox.button("Próxima Página ➡️", disabled=(st.session_state.pagina_ser >= total_paginas_ser), use_container_width=True, key="btn_prox_ser"):
                st.session_state.pagina_ser += 1; st.rerun()

    elif st.session_state.sub_tela_ser == "formulario":
        modo = st.session_state.modo_form_ser
        dados = st.session_state.dados_sel_ser
        campos_bloqueados = True if modo in ["visualizar", "excluir"] else False
        
        if modo == "cadastro": st.subheader("➕ Cadastrar Novo Serviço")
        elif modo == "visualizar": st.subheader(f"👁️ Visualizando Registro ID: {dados['ID']}")
        elif modo == "editar": st.subheader(f"✏️ Editando Registro ID: {dados['ID']}")
        elif modo == "excluir":
            st.subheader(f"⚠️ Confirmar Exclusão do Registro ID: {dados['ID']}")
            st.error(f"Atenção: Você está prestes a deletar o serviço '{dados['Nome']}'. Esta ação não pode ser desfeita.")

        novo_nome = st.text_input("Nome do Serviço", value=dados["Nome"], disabled=campos_bloqueados, key="input_ser_nome").strip()
        
        lista_tipo = ["SROP", "Gemini"]
        idx_tipo = lista_tipo.index(dados.get("Tipo", "SROP")) if dados.get("Tipo", "SROP") in lista_tipo else 0
        novo_tipo = st.selectbox("Tipo", lista_tipo, index=idx_tipo, disabled=campos_bloqueados, key="input_ser_tipo")
        
        lista_status = ["Ativo", "Inativo"]
        idx_status = lista_status.index(dados.get("Status", "Ativo")) if dados.get("Status", "Ativo") in lista_status else 0
        novo_status = st.selectbox("Situação", lista_status, index=idx_status, disabled=campos_bloqueados, key="input_ser_status")
        
        lista_exibir = ["Sim", "Não"]
        idx_exibir = lista_exibir.index(dados.get("Exibir_No_Menu", "Sim")) if dados.get("Exibir_No_Menu", "Sim") in lista_exibir else 0
        novo_exibir_menu = st.selectbox("Exibir no Menu", lista_exibir, index=idx_exibir, disabled=campos_bloqueados, key="input_ser_exibir")
        
        if novo_tipo == "SROP":
            nova_url_login = st.text_input("Endereço da Tela de Login", value=dados["UrlLogin"], disabled=campos_bloqueados, key="input_ser_urllogin").strip()
            nova_url_consulta = st.text_input("Endereço da Tela de Consulta", value=dados["UrlConsulta"], disabled=campos_bloqueados, key="input_ser_urlconsulta").strip()
            nova_url_pdf = st.text_input("Endereço de Extração do PDF", value=dados["UrlPdf"], disabled=campos_bloqueados, key="input_ser_urlpdf").strip()
            novo_login = st.text_input("Login", value=dados["Login"], disabled=campos_bloqueados, key="input_ser_login").strip()
            nova_senha = st.text_input("Senha", value=dados["Senha"], type="password", disabled=campos_bloqueados, key="input_ser_senha").strip()
            
            lista_dupla = ["Não", "Sim"]
            idx_dupla = lista_dupla.index(dados.get("DuplaAutenticacao", "Não")) if dados.get("DuplaAutenticacao", "Não") in lista_dupla else 0
            nova_dupla = st.selectbox("Dupla Autenticação", lista_dupla, index=idx_dupla, disabled=campos_bloqueados, key="input_ser_dupla")
            
            st.write("")
            novo_tempo_expiracao = st.number_input("Tempo Máximo de Sessão (Horas)", min_value=1, max_value=24, value=dados.get("Tempo_Expiracao_Horas", 4), disabled=campos_bloqueados, key="input_ser_tempo")
            
            st.write("")
            df_layouts = db.listar_layouts()
            lista_layouts = df_layouts["Nome_Layout"].tolist() if not df_layouts.empty else ["Layout Genérico Padrão"]
            id_atual_layout = dados.get("Layout_ID", 1)
            nome_atual_layout = df_layouts[df_layouts["ID"] == id_atual_layout]["Nome_Layout"].values[0] if not df_layouts.empty and id_atual_layout in df_layouts["ID"].values else lista_layouts[0]
            idx_layout = lista_layouts.index(nome_atual_layout) if nome_atual_layout in lista_layouts else 0
            
            novo_layout_nome = st.selectbox("Layout de Extração (OCR)", lista_layouts, index=idx_layout, disabled=campos_bloqueados, key="input_ser_layout")
            novo_layout_id = int(df_layouts[df_layouts["Nome_Layout"] == novo_layout_nome]["ID"].values[0]) if not df_layouts.empty else 1
            
        elif novo_tipo == "Gemini":
            if modo in ["visualizar", "excluir"] and dados.get("Senha"):
                valor_exibicao = "*** CREDENCIAIS SALVAS E CRIPTOGRAFADAS NO BANCO ***"
            elif modo == "editar" and dados.get("Senha"):
                valor_exibicao = "*** CREDENCIAIS SALVAS E CRIPTOGRAFADAS NO BANCO ***\n\n(Para manter as credenciais atuais, não altere este campo.\nPara substituí-las, apague todo este texto e cole o novo arquivo JSON aqui.)"
            else:
                valor_exibicao = dados.get("Senha", "")
                
            nova_senha_input = st.text_area("JSON (Credenciais)", value=valor_exibicao, disabled=campos_bloqueados, key="input_ser_json", height=200).strip()
            
            # Se o usuário não alterou o texto de exibição, mantemos a senha original (já descriptografada na memória)
            if "*** CREDENCIAIS SALVAS E CRIPTOGRAFADAS NO BANCO ***" in nova_senha_input:
                nova_senha = dados.get("Senha", "")
            else:
                nova_senha = nova_senha_input
                
            nova_url_login = "-"
            nova_url_consulta = "-"
            nova_url_pdf = "-"
            novo_login = "-"
            nova_dupla = "Não"
            novo_tempo_expiracao = 4
            novo_layout_id = 1
        
        st.write("")
        if modo == "cadastro":
            c_salvar, c_cancelar = st.columns(2)
            if c_salvar.button("💾 Salvar no Banco", key="btn_salvar_ser", width="stretch"):
                valido = bool(novo_nome and nova_url_login and nova_url_consulta and nova_url_pdf) if novo_tipo == "SROP" else bool(novo_nome and nova_senha)
                if not valido:
                    st.error("Por favor, preencha todos os campos obrigatórios do formulário.")
                else:
                    liberado = True
                    if novo_tipo == "SROP":
                        tags_obrigatorias = ["{DataInicialRegistro}", "{DataFinalRegistro}", "{idMunicipio}", "{size}"]
                        tags_faltantes = [t for t in tags_obrigatorias if t not in nova_url_consulta]
                        if tags_faltantes:
                            st.error(f"⚠️ Erro: O Endereço da Tela de Consulta deve conter todas as tags obrigatórias do SROP. Tags ausentes: {', '.join(tags_faltantes)}")
                            liberado = False
                    
                    if liberado:
                        sucesso = db.salvar_registro("servicos", {
                            "Nome": novo_nome, "UrlLogin": nova_url_login, "UrlConsulta": nova_url_consulta,
                            "UrlPdf": nova_url_pdf, "Login": novo_login, "Senha": nova_senha,
                            "DuplaAutenticacao": nova_dupla, "Tipo": novo_tipo, "Status": novo_status,
                            "Tempo_Expiracao_Horas": novo_tempo_expiracao, "Layout_ID": novo_layout_id,
                            "Exibir_No_Menu": novo_exibir_menu
                        })
                        if sucesso:
                            st.session_state.sub_tela_ser = "listar"
                            st.session_state.mensagem_sucesso = f"🎉 Serviço '{novo_nome}' cadastrado com sucesso!"
                            st.rerun()
                        else: st.error(f"⚠️ Erro: O serviço '{novo_nome}' já existe!")
            if c_cancelar.button("❌ Cancelar e Voltar", key="btn_cancelar_cad_ser", width="stretch"):
                st.session_state.sub_tela_ser = "listar"; st.rerun()
                    
        elif modo == "visualizar":
            if st.button("⬅️ Voltar para a Consulta", key="btn_voltar_vis_ser", width="stretch"):
                st.session_state.sub_tela_ser = "listar"; st.rerun()
                
        elif modo == "editar":
            c_atualizar, c_cancelar = st.columns(2)
            if c_atualizar.button("💾 Salvar Alterações", key="btn_update_ser", width="stretch"):
                valido = bool(novo_nome and nova_url_login and nova_url_consulta and nova_url_pdf) if novo_tipo == "SROP" else bool(novo_nome and nova_senha)
                if not valido:
                    st.error("Por favor, preencha todos os campos obrigatórios do formulário.")
                else:
                    liberado = True
                    if novo_tipo == "SROP":
                        tags_obrigatorias = ["{DataInicialRegistro}", "{DataFinalRegistro}", "{idMunicipio}", "{size}"]
                        tags_faltantes = [t for t in tags_obrigatorias if t not in nova_url_consulta]
                        if tags_faltantes:
                            st.error(f"⚠️ Erro: O Endereço da Tela de Consulta deve conter todas as tags obrigatórias do SROP. Tags ausentes: {', '.join(tags_faltantes)}")
                            liberado = False
                    
                    if liberado:
                        conn = db.obter_conexao()
                        cursor = conn.cursor()
                        senha_criptografada = db.criptografar_senha(nova_senha)
                        cursor.execute(
                            "UPDATE servicos SET Nome = %s, UrlLogin = %s, UrlConsulta = %s, UrlPdf = %s, Login = %s, Senha = %s, DuplaAutenticacao = %s, Tipo = %s, Status = %s, Tempo_Expiracao_Horas = %s, Layout_ID = %s, Exibir_No_Menu = %s WHERE ID = %s",
                            (novo_nome, nova_url_login, nova_url_consulta, nova_url_pdf, novo_login, senha_criptografada, nova_dupla, novo_tipo, novo_status, novo_tempo_expiracao, novo_layout_id, novo_exibir_menu, dados["ID"])
                        )
                        conn.commit()
                        conn.close()
                        st.session_state.sub_tela_ser = "listar"
                        st.session_state.mensagem_sucesso = "✏️ Serviço alterado com sucesso!"
                        st.rerun()
            if c_cancelar.button("❌ Cancelar", key="btn_cancelar_edit_ser", width="stretch"):
                st.session_state.sub_tela_ser = "listar"; st.rerun()
                
        elif modo == "excluir":
            c_deletar, c_voltar = st.columns(2)
            if c_deletar.button("🗑️ Sim, Excluir Registro", key="btn_delete_confirm_ser", width="stretch"):
                db.excluir_registro("servicos", dados["ID"])
                st.session_state.sub_tela_ser = "listar"
                st.session_state.mensagem_sucesso = f"🗑️ Serviço '{dados['Nome']}' removido!"
                st.rerun()
            if c_voltar.button("❌ Cancelar e Manter", key="btn_voltar_del_ser", width="stretch"):
                st.session_state.sub_tela_ser = "listar"; st.rerun()

    elif st.session_state.sub_tela_ser == "monitoramento":
        st.subheader("🖥️ Monitoramento Global de Sessões")
        
        col_voltar, col_limpar = st.columns(2)
        if col_voltar.button("⬅️ Voltar para Manutenção de Serviços", width="stretch"):
            st.session_state.sub_tela_ser = "listar"
            st.rerun()
        if col_limpar.button("🧹 Limpar Histórico de Inativos", width="stretch"):
            db.limpar_historico_inativo()
            st.session_state.mensagem_sucesso = "Todo o histórico de sessões antigas foi apagado com sucesso."
            st.rerun()
            
        st.write("")
        df_sessoes = db.listar_dados("servicos_sessoes")
        if df_sessoes.empty:
            st.info("🟢 Nenhuma sessão pendente. O sistema está limpo.")
        else:
            df_servicos = db.listar_dados("servicos")
            if not df_servicos.empty:
                df_sessoes = df_sessoes.merge(df_servicos[['ID', 'Nome']], left_on='Servico_ID', right_on='ID', how='left', suffixes=('', '_servico'))
                df_sessoes = df_sessoes.rename(columns={'Nome': 'Serviço'})
                
            # Ordena da sessão mais recente para a mais antiga (Ordem Decrescente)
            df_sessoes = df_sessoes.sort_values(by='Data_Login', ascending=False)
                
            from datetime import datetime
            
            for idx, row in df_sessoes.iterrows():
                id_sessao_db = row['ID']
                id_ser = row['Servico_ID']
                nome_ser = row.get('Serviço', f'Desconhecido (ID {id_ser})')
                data_login_str = row['Data_Login']
                status_sessao = row.get('Status', 'Ativa')
                
                try:
                    dt_login = datetime.strptime(data_login_str, '%Y-%m-%d %H:%M:%S')
                    logado_em_fmt = dt_login.strftime('%d/%m/%Y %H:%M:%S')
                    
                    if status_sessao == 'Ativa':
                        diff = datetime.now() - dt_login
                        total_seconds = int(diff.total_seconds())
                        hours = total_seconds // 3600
                        minutes = (total_seconds % 3600) // 60
                        seconds = total_seconds % 60
                        tempo_ativo_fmt = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                    else:
                        tempo_ativo_fmt = "--:--:--"
                except:
                    logado_em_fmt = data_login_str
                    tempo_ativo_fmt = "?"

                if status_sessao == "Ativa":
                    icon = "🟢"
                    status_lbl = f"<span style='color: #166534;'>{status_sessao.upper()}</span>"
                else:
                    icon = "⚪"
                    status_lbl = f"<span style='color: #6c757d;'>{status_sessao.upper()}</span>"
                
                with st.container():
                    col_status, col_servico, col_logado, col_tempo, col_btn = st.columns([1.0, 2.0, 2.5, 1.5, 3.0])
                    
                    with col_status:
                        st.markdown(f"<div style='padding-top: 5px; font-size: 12px; color: #6c757d; text-transform: uppercase;'>Status</div><div style='padding-top: 2px; font-size: 15px;'><b>{icon} {status_lbl}</b></div>", unsafe_allow_html=True)
                    with col_servico:
                        st.markdown(f"<div style='padding-top: 5px; font-size: 12px; color: #6c757d; text-transform: uppercase;'>Serviço</div><div style='padding-top: 2px; font-size: 15px;'><b style='color: #343a40;'>{nome_ser}</b></div>", unsafe_allow_html=True)
                    with col_logado:
                        st.markdown(f"<div style='padding-top: 5px; font-size: 12px; color: #6c757d; text-transform: uppercase;'>Logado em</div><div style='padding-top: 2px; font-size: 15px;'><b style='color: #343a40;'>{logado_em_fmt}</b></div>", unsafe_allow_html=True)
                    with col_tempo:
                        st.markdown(f"<div style='padding-top: 5px; font-size: 12px; color: #6c757d; text-transform: uppercase;'>Tempo Ativo</div><div style='padding-top: 2px; font-size: 15px;'><b style='color: #343a40;'>{tempo_ativo_fmt}</b></div>", unsafe_allow_html=True)
                    
                    with col_btn:
                        st.markdown("<div style='height: 12px;'></div>", unsafe_allow_html=True)
                        if status_sessao == "Ativa":
                            cb_test, cb_kill = st.columns(2)
                            if cb_test.button("🔄 Testar", key=f"btn_test_sess_{id_sessao_db}", use_container_width=True):
                                with st.spinner("Ping..."):
                                    import automacao as aut
                                    import json
                                    
                                    # Pega URL de Consulta do serviço
                                    df_ser_url = db.listar_dados("servicos")
                                    url_consulta = df_ser_url[df_ser_url["ID"] == id_ser]["UrlConsulta"].iloc[0] if not df_ser_url.empty else ""
                                    
                                    # Pega os cookies
                                    sess_data_str = db.descriptografar_senha(row['Session_Data'])
                                    if sess_data_str != "Erro ao descriptografar":
                                        sess_obj = json.loads(sess_data_str)
                                        is_alive = aut.testar_conexao_srop(sess_obj, url_consulta)
                                        if is_alive:
                                            st.toast("✅ A sessão respondeu e está perfeitamente viva no SROP!", icon="✅")
                                        else:
                                            db.atualizar_status_sessao(id_sessao_db, "Expirada (Remota)")
                                            st.error("❌ Sessão morreu lá no servidor. Atualizando a grade...")
                                            st.rerun()
                                    else:
                                        st.error("Erro de descriptografia")
                                        
                            if cb_kill.button("🔴 Encerrar", key=f"btn_kill_sess_{id_sessao_db}", use_container_width=True):
                                db.limpar_sessao(id_ser)
                                st.session_state.mensagem_sucesso = f"A sessão '{nome_ser}' foi encerrada com sucesso."
                                st.rerun()
                        else:
                            if st.button("🗑️ Apagar", key=f"btn_kill_sess_{id_sessao_db}", use_container_width=True):
                                db.excluir_historico_sessao(id_sessao_db)
                                st.session_state.mensagem_sucesso = f"O registro de histórico foi apagado permanentemente."
                                st.rerun()
                                
                    st.markdown("<hr style='margin: 15px 0; border: none; border-top: 1px solid #e9ecef;'>", unsafe_allow_html=True)


# =====================================================================
# TELA: MANUTENÇÃO DE LAYOUTS
# =====================================================================
elif menu == "📄 Manutenção de Layouts":
    st.title("📄 Manutenção de Layouts (OCR Dinâmico)")
    st.info("Crie moldes dinâmicos para a extração inteligente de dados dos Boletins de Ocorrência.")
    
    # 1. Escolher ou Criar Layout
    col_l1, col_l2, col_l3 = st.columns([3, 1, 1])
    layouts_df = db.listar_layouts()
    opcoes_layout = ["-- Selecione um Layout --", "➕ Criar Novo Layout..."]
    
    if not layouts_df.empty:
        opcoes_layout.extend(layouts_df["Nome_Layout"].tolist())
        
    idx_default = 2 if len(layouts_df) == 1 else 0
    layout_selecionado = col_l1.selectbox("Selecione o Layout", opcoes_layout, index=idx_default)
    
    if layout_selecionado == "➕ Criar Novo Layout...":
        with st.form("form_novo_layout"):
            novo_nome = st.text_input("Nome do Novo Layout")
            if st.form_submit_button("Salvar Layout"):
                if db.salvar_layout(novo_nome):
                    st.success("Layout criado com sucesso!")
                    st.rerun()
                else:
                    st.error("Erro ao salvar ou Layout já existe.")
    
    elif layout_selecionado != "-- Selecione um Layout --":
        # Pega o ID do layout selecionado
        layout_id = int(layouts_df[layouts_df["Nome_Layout"] == layout_selecionado]["ID"].values[0])
        
        # Botão de exclusão do layout
        with col_l2:
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            if st.button("🗑️ Excluir Layout Inteiro", use_container_width=True):
                if db.excluir_layout(layout_id):
                    st.success("Layout excluído!")
                    st.rerun()
                else:
                    st.error("Não é possível excluir este layout (está em uso por algum serviço).")
                    
        # Popover para clonar o layout selecionado
        with col_l3:
            st.markdown("<div style='margin-top: 28px;'></div>", unsafe_allow_html=True)
            with st.popover("👥 Clonar Layout", use_container_width=True):
                st.markdown(f"**Clonar Layout:** {layout_selecionado}")
                nome_sugerido = f"{layout_selecionado} (Cópia)"
                novo_nome_clone = st.text_input("Nome do novo Layout", value=nome_sugerido, key="input_clone_layout_nome")
                if st.button("Confirmar Clone", use_container_width=True):
                    if not novo_nome_clone.strip():
                        st.error("O nome não pode ser vazio.")
                    elif db.clonar_layout(layout_id, novo_nome_clone.strip()):
                        st.success("Layout clonado com sucesso!")
                        st.rerun()
                    else:
                        st.error("Erro ao clonar o layout. Verifique se o nome já existe.")
                    
        # Central de Ajuda de Chaves Especiais e Fórmulas
        import extrair_bo as ex_bo
        with st.expander("💡 Central de Ajuda: Fórmulas e Metadados Especiais", expanded=False):
            # Calcular exemplos para hora estática
            hora_ex_padrao = "14:30"
            faixa_ex = ex_bo.formatar_faixa_horario(hora_ex_padrao)
            periodo_ex = ex_bo.formatar_periodo_dia(hora_ex_padrao)
            
            st.markdown(f"""
            Você pode cadastrar itens especiais no layout usando palavras-chave específicas no campo **Palavra Busca (PDF)**:
            
            ### 📂 Metadados do Arquivo e do BO:
            * `*ARQUIVO*` -> Retorna o nome do arquivo PDF (ex: `bo_12345.pdf`).
            * `*BO_NUMERO*` -> Retorna o número do Boletim de Ocorrência.
            * `*DATA_DO_REGISTRO*` -> Retorna a data em que o BO foi registrado.
            * `*HORA_DO_REGISTRO*` -> Retorna o horário de registro do BO.
            
            ### 🔢 Fórmulas de Data e Hora Dinâmicas:
            Você pode extrair parcelas de **qualquer campo de data/hora** do BO (ex: a partir de `DATA_DO_FATO`, `HORA_DO_FATO`, etc.) usando a sintaxe:
            * `*DIA:CAMPO_ORIGEM*` -> Extrai o dia com 2 dígitos. Exemplo: `*DIA:DATA_DO_FATO*`
            * `*MES:CAMPO_ORIGEM*` -> Extrai o mês com 2 dígitos. Exemplo: `*MES:DATA_DO_FATO*`
            * `*ANO:CAMPO_ORIGEM*` -> Extrai o ano com 4 dígitos. Exemplo: `*ANO:DATA_DO_FATO*`
            * `*DIA_SEMANA:CAMPO_ORIGEM*` -> Extrai o dia da semana por extenso em maiúsculo. Exemplo: `*DIA_SEMANA:DATA_DO_FATO*` (Gera `SEGUNDA-FEIRA`, `TERÇA-FEIRA`, etc.)
            * `*HORA:CAMPO_ORIGEM*` -> Extrai apenas a hora (HH) com 2 dígitos de um campo de hora. Exemplo: `*HORA:HORA_DO_FATO*` (Extrai `14` a partir de `14:30`)
            * `*FAIXA_HORA:CAMPO_ORIGEM*` -> Analisa a hora (HH:MM) e retorna a faixa de 3 horas correspondente. Exemplo: `*FAIXA_HORA:HORA_DO_FATO*` (Gera `{faixa_ex}` para `14:30`)
            * `*PERIODO_HORA:CAMPO_ORIGEM*` -> Analisa a hora (HH:MM) e retorna o período do dia (MADRUGADA, MATUTINO, VESPERTINO, NOTURNO). Exemplo: `*PERIODO_HORA:HORA_DO_FATO*` (Gera `{periodo_ex}` para `14:30`)
            
            ### 🔗 Fórmulas de Concatenação de Colunas:
            Você pode juntar valores de múltiplas colunas existentes do layout em uma única célula separados por vírgula:
            * `*CONCAT:CAMPO1,CAMPO2,CAMPO3,...*` -> Concatena as colunas listadas separando-as por vírgula e espaço (campos vazios são ignorados de forma inteligente). Exemplo: `*CONCAT:BAIRRO,MUNICIPIO,SIGL ESTADO*` (Gera `"Centro, São Paulo, SP"`).
            
            ### 📌 Fórmulas de Valores Fixos/Constantes:
            Você pode definir colunas com valores fixos e constantes na planilha:
            * `*FIXO:valor*` -> Retorna o valor fixo digitado. Exemplos: `*FIXO:1*` (gera o número 1), `*FIXO:Sim*` (gera a palavra "Sim"), `*FIXO:*` (gera uma célula vazia).
            
            *Dica: Você pode mudar livremente o nome da coluna de destino e a ordem de exibição delas no Excel!*
            """)
            
            st.markdown("---")
            st.markdown("#### ⏱️ Testador Interativo de Horários")
            st.write("Digite um horário no formato HH:MM para simular as fórmulas na interface:")
            
            hora_atual_padrao = datetime.now().strftime("%H:%M")
            col_t1, col_t2, col_t3 = st.columns([2, 3, 3])
            
            hora_teste_in = col_t1.text_input("Horário (HH:MM)", value=hora_atual_padrao, key="input_hora_teste_s")
            
            faixa_resultado = ex_bo.formatar_faixa_horario(hora_teste_in)
            periodo_resultado = ex_bo.formatar_periodo_dia(hora_teste_in)
            
            if faixa_resultado:
                col_t2.info(f"**Faixa de Horário Calculada:**\n`{faixa_resultado}`")
            else:
                col_t2.error("**Faixa de Horário:**\nHorário inválido ou incorreto")
                
            if periodo_resultado:
                col_t3.info(f"**Período do Dia Calculado:**\n`{periodo_resultado}`")
            else:
                col_t3.error("**Período do Dia:**\nHorário inválido ou incorreto")
        
        st.markdown("---")
        
        grupos_df = db.listar_grupos(layout_id)
        
        # 2. Dashboard Header (Metrics)
        m_col1, m_col2, m_col3 = st.columns(3)
        m_col1.metric("Total de Grupos", len(grupos_df))
        
        total_itens = 0
        if not grupos_df.empty:
            for _, r in grupos_df.iterrows():
                if r["Tem_Itens"]:
                    total_itens += len(db.listar_itens(r["ID"]))
        m_col2.metric("Total de Itens", total_itens)
        
        st.write("")
        
        # 3. Formulário para novo grupo
        with st.expander("➕ Adicionar Novo Grupo", expanded=grupos_df.empty):
            with st.form("form_novo_grupo", clear_on_submit=True):
                g_col1, g_col2a, g_col2b, g_col3, g_col4 = st.columns([2, 1, 1, 1, 1])
                novo_grupo = g_col1.text_input("Nome do Grupo (Ex: VÍTIMA)")
                ordem_grupo = g_col2a.number_input("Ordem Leitura (Pastas)", min_value=1, value=1)
                ordem_excel_g = g_col2b.number_input("Ordem Excel (Colunas)", min_value=1, value=1)
                tem_itens = g_col3.checkbox("Tem Sub-itens?", value=True, help="Desmarque se o grupo for apenas um texto direto sem chave-valor.")
                exportar_excel_g = g_col4.checkbox("Exportar Excel?", value=True, help="Se desmarcado, este grupo não será gerado como coluna no Excel final.")
                
                if st.form_submit_button("Salvar Grupo"):
                    if db.salvar_grupo(layout_id, novo_grupo, ordem_grupo, 1 if tem_itens else 0, 1 if exportar_excel_g else 0, ordem_excel_g):
                        st.success("Grupo adicionado!")
                        st.rerun()
                    else:
                        st.error("Erro ao salvar ou Grupo já existe neste layout.")
                        
        st.write("")
        
        # 4. Renderiza os grupos em Expanders (Acordeão)
        if not grupos_df.empty:
            for index, row_g in grupos_df.iterrows():
                grupo_id = row_g["ID"]
                grupo_nome = row_g["Nome_Grupo"]
                grupo_ordem = row_g["Ordem"]
                grupo_tem_itens = row_g["Tem_Itens"]
                grupo_exportar = row_g.get("Exportar_Excel", 1)
                
                # Cada grupo é um card colapsável com indicação visual de exportação
                label_prefix = "📁" if grupo_exportar == 1 else "📁 🚫"
                label_suffix = "" if grupo_exportar == 1 else " (Não Exportar)"
                with st.expander(f"{label_prefix} {grupo_ordem}. {grupo_nome}{label_suffix}", expanded=False):
                    
                    editando_g = st.session_state.get(f"edit_g_{grupo_id}", False)
                    if editando_g:
                        st.markdown("**Editar Grupo:**")
                        ge_col1, ge_col2a, ge_col2b, ge_col3, ge_col4, ge_col5 = st.columns([2, 1, 1, 1, 1, 1])
                        g_new_nome = ge_col1.text_input("Nome Grupo", value=grupo_nome, key=f"gn_{grupo_id}")
                        g_new_ordem = ge_col2a.number_input("Ordem Leitura", value=int(grupo_ordem), min_value=1, key=f"go_{grupo_id}")
                        g_new_ordem_excel = ge_col2b.number_input("Ordem Excel", value=int(row_g.get("Ordem_Excel", grupo_ordem)), min_value=1, key=f"goe_{grupo_id}")
                        g_new_tem_itens = ge_col3.checkbox("Tem Sub-itens?", value=bool(grupo_tem_itens), key=f"gt_{grupo_id}")
                        g_new_exportar = ge_col4.checkbox("Exportar Excel?", value=bool(grupo_exportar == 1), key=f"gex_{grupo_id}")
                        
                        ge_btn1, ge_btn2 = ge_col5.columns(2)
                        st.write("") # spacer
                        if ge_btn1.button("💾 Salvar", key=f"gsv_{grupo_id}", help="Salvar Grupo"):
                            if db.atualizar_grupo(grupo_id, layout_id, g_new_nome, g_new_ordem, 1 if g_new_tem_itens else 0, 1 if g_new_exportar else 0, g_new_ordem_excel):
                                st.session_state[f"edit_g_{grupo_id}"] = False
                                st.rerun()
                            else:
                                st.error("Erro ou Nome já existe!")
                        if ge_btn2.button("❌ Cancelar", key=f"gcc_{grupo_id}", help="Cancelar Edição"):
                            st.session_state[f"edit_g_{grupo_id}"] = False
                            st.rerun()
                        st.write("")
                    else:
                        _, col_edt_g, col_reset_g, col_del_g = st.columns([4.5, 1.5, 2.0, 2.0])
                        if col_edt_g.button("✏️ Editar Grupo", key=f"edt_g_{grupo_id}", use_container_width=True):
                            st.session_state[f"edit_g_{grupo_id}"] = True
                            st.rerun()
                        if col_reset_g.button("🔄 Resetar Ordenação", key=f"reset_g_{grupo_id}", use_container_width=True, help="Reseta os valores de ordenação de todos os itens do grupo para 1"):
                            if db.resetar_ordenacao_grupo(grupo_id):
                                st.success("Ordenação resetada com sucesso!")
                                st.rerun()
                            else:
                                st.error("Erro ao resetar ordenação do grupo.")
                        if col_del_g.button(f"🗑️ Excluir Grupo Inteiro", key=f"del_g_{grupo_id}", use_container_width=True):
                            db.excluir_grupo(grupo_id)
                            st.rerun()
                        st.write("")
                    
                    if grupo_tem_itens:
                        itens_df = db.listar_itens(grupo_id)
                        
                        # Formulario de novo item escondido dentro de um expander secundário
                        with st.expander("➕ Adicionar Novo Item", expanded=itens_df.empty):
                            with st.form(f"form_item_{grupo_id}", clear_on_submit=True):
                                i_col1, i_col2, i_col3a, i_col3b, i_col4 = st.columns([2, 2, 1, 1, 1])
                                nome_excel = i_col1.text_input("Coluna Excel (Destino)", help="Nome da coluna que será gerada no Excel")
                                palavra_busca = i_col2.text_input("Palavra Busca (PDF)", help="Dica: Use a barra | para diferenciar palavras repetidas. Exemplo: UF|Naturalidade")
                                ordem_item = i_col3a.number_input("Ordem Leitura", min_value=1, value=1, help="Ordem de extração deste item no PDF")
                                ordem_excel_item = i_col3b.number_input("Ordem Excel", min_value=1, value=1, help="Ordem da coluna no Excel")
                                
                                i_col4.markdown("<div style='margin-top: 36px;'></div>", unsafe_allow_html=True)
                                exportar_excel_bool = i_col4.checkbox("Exportar para Excel", value=True, help="Se desmarcado, este item servirá apenas como barreira de busca, mas não será gerado como coluna no arquivo Excel final.")
                                
                                if st.form_submit_button("Salvar Item"):
                                    exportar_excel = 1 if exportar_excel_bool else 0
                                    if db.salvar_item(layout_id, grupo_id, nome_excel, palavra_busca, ordem_item, exportar_excel, ordem_excel_item):
                                        st.success("Item adicionado!")
                                        st.rerun()
                                    else:
                                        st.error("Nome da coluna já existe neste layout ou campos inválidos!")
                        
                        # Grid de exibição dos itens (Tabela simulada com colunas)
                        if not itens_df.empty:
                            st.write("") # Espaçamento top
                            c_title, c_bt1, c_bt2 = st.columns([5, 2.5, 2.5])
                            c_title.markdown("**Itens de Extração Cadastrados:**")
                            
                            def toggle_all_items(gid, state):
                                df_items = db.listar_itens(gid)
                                for _, r in df_items.iterrows():
                                    iid = int(r['ID'])
                                    db.atualizar_exportacao_item(iid, state)
                                    st_key = f"q_nx_{iid}"
                                    if st_key in st.session_state:
                                        st.session_state[st_key] = bool(state)
                                    
                            if c_bt1.button("✅ Marcar Todos", key=f"marcar_tds_{grupo_id}", use_container_width=True, on_click=toggle_all_items, args=(grupo_id, 1)):
                                pass
                                
                            if c_bt2.button("❌ Desmarcar Todos", key=f"desmarcar_tds_{grupo_id}", use_container_width=True, on_click=toggle_all_items, args=(grupo_id, 0)):
                                pass
                            
                            st.write("") # Espaçamento bot
                            
                            # Cabeçalhos
                            h_col1a, h_col1b, h_col2, h_col3, h_colE, h_col4 = st.columns([0.7, 0.7, 2, 2, 1.2, 0.8])
                            h_col1a.markdown("**Ordem L.**")
                            h_col1b.markdown("**Ordem E.**")
                            h_col2.markdown("**Coluna Excel**")
                            h_col3.markdown("**Palavra Busca (PDF)**")
                            h_colE.markdown("**Exportar no Excel**")
                            h_col4.markdown("**Ação**")
                            st.markdown("<hr style='margin: 5px 0; padding: 0;'>", unsafe_allow_html=True)
                            
                            # Linhas
                            for _, row_i in itens_df.iterrows():
                                item_id = row_i['ID']
                                editando = st.session_state.get(f"edit_i_{item_id}", False)
                                
                                if editando:
                                    e_col1a, e_col1b, e_col2, e_col3, e_colE, e_col4, e_col5 = st.columns([0.7, 0.7, 2, 2, 1.2, 0.4, 0.4])
                                    n_ordem = e_col1a.number_input("Ordem L", value=int(row_i['Ordem']), min_value=1, key=f"no_{item_id}", label_visibility="collapsed")
                                    n_ordem_excel = e_col1b.number_input("Ordem E", value=int(row_i.get('Ordem_Excel', row_i['Ordem'])), min_value=1, key=f"noe_{item_id}", label_visibility="collapsed")
                                    n_excel = e_col2.text_input("Excel", value=row_i['Nome_Item_Excel'], key=f"ne_{item_id}", label_visibility="collapsed")
                                    n_busca = e_col3.text_input("Busca", value=row_i['Palavra_Busca'], key=f"nb_{item_id}", label_visibility="collapsed", help="Dica: Use a barra | para diferenciar palavras repetidas.")
                                    n_export = e_colE.checkbox("Sim", value=bool(row_i.get('Exportar_Excel', 1)), key=f"nx_{item_id}")
                                    
                                    if e_col4.button("💾", key=f"sv_{item_id}", help="Salvar Edição"):
                                        if db.atualizar_item(item_id, layout_id, n_excel, n_busca, n_ordem, 1 if n_export else 0, n_ordem_excel):
                                            st.session_state[f"edit_i_{item_id}"] = False
                                            st.rerun()
                                        else:
                                            st.error("Erro ou Coluna já existe!")
                                            
                                    if e_col5.button("❌", key=f"cc_{item_id}", help="Cancelar Edição"):
                                        st.session_state[f"edit_i_{item_id}"] = False
                                        st.rerun()
                                else:
                                    r_col1a, r_col1b, r_col2, r_col3, r_colE, r_col4, r_col5 = st.columns([0.7, 0.7, 2, 2, 1.2, 0.4, 0.4])
                                    r_col1a.write(f"#{row_i['Ordem']}")
                                    r_col1b.write(f"📊#{row_i.get('Ordem_Excel', row_i['Ordem'])}")
                                    r_col2.markdown(f"`{row_i['Nome_Item_Excel']}`")
                                    r_col3.markdown(f"`{row_i['Palavra_Busca']}`")
                                    def toggle_single_item(iid, st_key):
                                        new_val = st.session_state[st_key]
                                        db.atualizar_exportacao_item(int(iid), 1 if new_val else 0)
                                        
                                    is_export = bool(row_i.get('Exportar_Excel', 1))
                                    r_colE.checkbox("Exp", value=is_export, key=f"q_nx_{item_id}", label_visibility="collapsed", on_change=toggle_single_item, args=(item_id, f"q_nx_{item_id}"))
                                    if r_col4.button("✏️", key=f"bt_edt_i_{item_id}", help="Editar"):
                                        st.session_state[f"edit_i_{item_id}"] = True
                                        st.rerun()
                                    if r_col5.button("🗑️", key=f"del_i_{item_id}", help="Excluir"):
                                        db.excluir_item(item_id)
                                        st.rerun()
                                st.markdown("<hr style='margin: 2px 0; padding: 0; border-top: 1px dashed #eee;'>", unsafe_allow_html=True)
                        else:
                            st.info("Nenhum item cadastrado neste grupo. Adicione itens acima.")
                    else:
                        st.info(f"O robô extrairá todo o texto que estiver abaixo da palavra **{grupo_nome}** até o próximo grupo.")
        
        # --- PRÉVIA DO CABEÇALHO DO EXCEL ---
        st.write("")
        st.markdown("### 📊 Prévia do Cabeçalho da Planilha Excel")
        st.write("Esta é a ordem exata das colunas no relatório Excel com base nas configurações acima:")
        
        # Função auxiliar para extrair a lista ordenada de colunas do layout
        def obter_colunas_preview_layout(l_id: int) -> list[str]:
            conn_prv = db.obter_conexao()
            cursor_prv = conn_prv.cursor()
            # 1. Busca todos os itens ativos
            cursor_prv.execute("""
                SELECT i.Nome_Item_Excel, i.Ordem_Excel, i.Palavra_Busca, i.ID
                FROM layout_itens i 
                JOIN layout_grupos g ON i.Grupo_ID = g.ID 
                WHERE g.Layout_ID = %s AND i.Exportar_Excel = 1
            """, (l_id,))
            itens_db = cursor_prv.fetchall()
            
            # 2. Busca todos os grupos ativos sem sub-itens (colunas diretas)
            cursor_prv.execute("""
                SELECT Nome_Grupo, Ordem_Excel, '*TEXTO_BRUTO*', ID
                FROM layout_grupos
                WHERE Layout_ID = %s AND Tem_Itens = 0 AND Exportar_Excel = 1
            """, (l_id,))
            grupos_db = cursor_prv.fetchall()
            
            # 3. Busca se existem itens de metadados especiais mapeados no banco (ativos ou não) para evitar fallback indevido
            cursor_prv.execute("""
                SELECT i.Palavra_Busca 
                FROM layout_itens i 
                JOIN layout_grupos g ON i.Grupo_ID = g.ID 
                WHERE g.Layout_ID = %s 
                  AND i.Palavra_Busca IN ('*ARQUIVO*', '*BO_NUMERO*', '*DATA_DO_REGISTRO*', '*HORA_DO_REGISTRO*')
            """, (l_id,))
            especiais_mapeados = [r[0] for r in cursor_prv.fetchall()]
            conn_prv.close()
            
            colunas_candidatas = []
            import extrair_bo as ex_bo
            for nome, ordem_e, palavra, iid in itens_db:
                colunas_candidatas.append({
                    "nome": ex_bo.normalizar_nome_chave(nome),
                    "nome_original": nome,
                    "ordem_excel": ordem_e,
                    "palavra": palavra,
                    "id": iid
                })
            for nome, ordem_e, palavra, iid in grupos_db:
                colunas_candidatas.append({
                    "nome": ex_bo.normalizar_nome_chave(nome),
                    "nome_original": nome,
                    "ordem_excel": ordem_e,
                    "palavra": palavra,
                    "id": iid
                })
                
            # Ordena pelo Ordem_Excel (crescente) e pelo ID (estável)
            colunas_candidatas.sort(key=lambda x: (x["ordem_excel"], x["id"]))
            
            mapeamento_exibicao = {}
            for c in colunas_candidatas:
                mapeamento_exibicao[c["nome"]] = c["nome_original"]
                
            headers = []
            # Adiciona metadados estáticos antigos apenas como fallback se não estiverem mapeados de forma alguma no banco (ativos ou inativos)
            if "*ARQUIVO*" not in especiais_mapeados:
                headers.append("ARQUIVO")
                mapeamento_exibicao["ARQUIVO"] = "ARQUIVO"
            if "*BO_NUMERO*" not in especiais_mapeados:
                headers.append("BO_NUMERO")
                mapeamento_exibicao["BO_NUMERO"] = "BO_NUMERO"
            if "*DATA_DO_REGISTRO*" not in especiais_mapeados:
                headers.append("DATA_DO_REGISTRO")
                mapeamento_exibicao["DATA_DO_REGISTRO"] = "DATA_DO_REGISTRO"
            if "*HORA_DO_REGISTRO*" not in especiais_mapeados:
                headers.append("HORA_DO_REGISTRO")
                mapeamento_exibicao["HORA_DO_REGISTRO"] = "HORA_DO_REGISTRO"
                
            for c in colunas_candidatas:
                nome_norm = c["nome"]
                if nome_norm not in headers:
                    headers.append(nome_norm)
                    
            # Converte para os nomes originais de exibição
            headers_exibicao = [mapeamento_exibicao.get(h, h) for h in headers]
            return headers_exibicao

        colunas_preview = obter_colunas_preview_layout(layout_id)
        if colunas_preview:
            # Mostra uma tabela vazia no Streamlit (uma linha com células vazias) simulando o Excel
            df_preview_mock = pd.DataFrame(columns=colunas_preview)
            st.dataframe(df_preview_mock, use_container_width=True)
            st.caption(f"Total de colunas ativas para exportação: **{len(colunas_preview)}**")
        else:
            st.info("Nenhuma coluna configurada para exportação neste layout ainda.")
# =====================================================================
# TELA: MANUTENÇÃO DE TIPO LOCAL (IA)
# =====================================================================
elif menu == "📍 Manutenção de Tipo Local":
    st.title("📍 Manutenção de Tipo Local")
    st.write("Cadastre os tipos de local e a descrição para que a Inteligência Artificial saiba como classificar.")
    
    with st.expander("➕ Adicionar Novo Tipo Local", expanded=False):
        with st.form("form_novo_tipo_local", clear_on_submit=True):
            col1, col2 = st.columns([1, 1])
            with col1:
                novo_tipo = st.text_input("Tipo Local (ex: VIA PÚBLICA)", max_chars=100)
            with col2:
                status_tipo = st.selectbox("Status", ["Ativo", "Inativo"])
                
            nova_desc = st.text_area("Descrição para a IA", placeholder="Ex: Quando o fato ocorrer em ruas, avenidas, calçadas...", height=100)
            
            if st.form_submit_button("💾 Salvar Tipo Local"):
                if not novo_tipo.strip() or not nova_desc.strip():
                    st.warning("⚠️ Preencha o Tipo Local e a Descrição.")
                else:
                    if db.salvar_tipo_local(novo_tipo, nova_desc, status_tipo):
                        st.success("✅ Tipo Local cadastrado com sucesso!")
                        st.rerun()
                    else:
                        st.error("⚠️ Erro ao cadastrar. Verifique se o Tipo Local já existe.")

    st.subheader("📋 Tipos de Local Cadastrados")
    df_tipos = db.listar_tipos_local()
    
    if df_tipos.empty:
        st.info("Nenhum tipo local cadastrado ainda.")
    else:
        # Cabeçalho da grid para melhor visualização
        st.markdown("<hr style='margin: 0px 0px 10px 0px;'>", unsafe_allow_html=True)
        col_h1, col_h2, col_h3, col_h4 = st.columns([3, 5, 2, 2])
        col_h1.markdown("**📌 Tipo Local**")
        col_h2.markdown("**📝 Descrição para IA**")
        col_h3.markdown("**Status**")
        col_h4.markdown("**Ações**")
        st.markdown("<hr style='margin: 10px 0px 10px 0px;'>", unsafe_allow_html=True)
        
        for index, row in df_tipos.iterrows():
            tipo_local = row.get('Tipo_Local', row.get('tipo_local', ''))
            status = row.get('Status', row.get('status', ''))
            desc_ia = row.get('Descricao_IA', row.get('descricao_ia', ''))
            row_id = row.get('ID', row.get('id', ''))
            
            c1, c2, c3, c4 = st.columns([3, 5, 2, 2])
            c1.markdown(f"**{tipo_local}**")
            c2.caption(desc_ia)
            if status == 'Ativo':
                c3.markdown("🟢 Ativo")
            else:
                c3.markdown("🔴 Inativo")
            
            with c4:
                cb1, cb2 = st.columns(2)
                edit_click = cb1.button("✏️", key=f"edit_tipo_{row_id}", help="Editar", use_container_width=True)
                del_click = cb2.button("🗑️", key=f"del_tipo_{row_id}", help="Excluir", use_container_width=True)
                
            if edit_click:
                st.session_state[f"editando_tipo_{row_id}"] = not st.session_state.get(f"editando_tipo_{row_id}", False)
            if del_click:
                db.excluir_tipo_local(row_id)
                st.rerun()
                
            if st.session_state.get(f"editando_tipo_{row_id}", False):
                with st.container():
                    with st.form(f"form_edit_tipo_{row_id}"):
                        st.write(f"**Editando:** {tipo_local}")
                        col_e1, col_e2 = st.columns([2, 1])
                        with col_e1:
                            edit_tipo = st.text_input("Tipo Local", value=tipo_local)
                        with col_e2:
                            edit_status = st.selectbox("Status", ["Ativo", "Inativo"], index=0 if status == 'Ativo' else 1)
                        edit_desc = st.text_area("Descrição", value=desc_ia, height=80)
                        
                        if st.form_submit_button("💾 Salvar Alterações"):
                            if db.atualizar_tipo_local(row_id, edit_tipo, edit_desc, edit_status):
                                st.success("Atualizado!")
                                st.session_state[f"editando_tipo_{row_id}"] = False
                                st.rerun()
                            else:
                                st.error("Erro ao atualizar. Nome duplicado?")
                                
            st.markdown("<hr style='margin: 5px 0px 10px 0px; border-top: 1px dashed #eee;'>", unsafe_allow_html=True)

# =====================================================================
# TELA: MANUTENÇÃO DE PROMPTS (IA)
# =====================================================================
elif menu == "📝 Manutenção de Prompts":
    st.title("📝 Manutenção de Prompts")
    st.write("Configure os prompts que a Inteligência Artificial utilizará para classificar os dados.")
    
    with st.expander("➕ Adicionar Novo Prompt", expanded=False):
        with st.form("form_novo_prompt", clear_on_submit=True):
            col1, col2 = st.columns([2, 1])
            with col1:
                novo_nome = st.text_input("Nome do Prompt", placeholder="Ex: Classificação de Tipo Local")
            with col2:
                novo_tipo_prompt = st.selectbox("Vincular ao Tipo", ["Tipo Local"]) # Futuro: adicionar mais
            
            st.write("Dica: Use `{CATEGORIAS}` onde a IA deve ler a lista de tipos, e `{TEXTO}` onde deve ler a narrativa.")
            nova_instrucao = st.text_area("Instrução para a IA", height=250,
                                          value="Você é um especialista em análise criminal e inteligência policial de ponta.\nSua ÚNICA função é analisar de forma fria e objetiva a NARRATIVA de um Boletim de Ocorrência Policial e classificar onde ocorreu o fato (TIPO LOCAL).\n\nAbaixo estão as ÚNICAS categorias permitidas e suas respectivas descrições:\n---\n{CATEGORIAS}\n---\n\nNARRATIVA DO BOLETIM PARA ANÁLISE:\n\"{TEXTO}\"\n\nREGRAS DE OURO ESTRITAS:\n1. O seu retorno deve ser APENAS o nome exato da categoria em letras maiúsculas. Absolutamente mais nada.\n2. Não forneça saudações, não justifique a escolha, não use aspas. Exemplo de resposta correta: VIA PÚBLICA\n3. Se a narrativa não descrever o local, for muito vaga, ou se o local não se encaixar com total clareza em NENHUMA das categorias acima, responda APENAS: NI\n4. Em caso de múltiplos locais na mesma narrativa (ex: começou na rua e terminou em casa), classifique o local onde o CRIME PRINCIPAL foi efetivamente consumado.")
            novo_status_prompt = st.selectbox("Status (Apenas 1 Ativo por Tipo)", ["Ativo", "Inativo"])
            
            if st.form_submit_button("💾 Salvar Prompt"):
                if not novo_nome.strip() or not nova_instrucao.strip():
                    st.warning("⚠️ Preencha o Nome e a Instrução.")
                else:
                    if db.salvar_prompt(novo_nome, novo_tipo_prompt, nova_instrucao, novo_status_prompt):
                        st.success("✅ Prompt cadastrado com sucesso!")
                        st.rerun()
                    else:
                        st.error("⚠️ Erro ao cadastrar.")

    st.subheader("📋 Prompts Cadastrados")
    df_prompts = db.listar_prompts()
    
    if df_prompts.empty:
        st.info("Nenhum prompt cadastrado ainda.")
    else:
        st.markdown("<hr style='margin: 0px 0px 10px 0px;'>", unsafe_allow_html=True)
        col_h1, col_h2, col_h3, col_h4 = st.columns([3, 2, 1, 2])
        col_h1.markdown("**📌 Nome do Prompt**")
        col_h2.markdown("**Vínculo**")
        col_h3.markdown("**Status**")
        col_h4.markdown("**Ações**")
        st.markdown("<hr style='margin: 10px 0px 10px 0px;'>", unsafe_allow_html=True)
        
        for index, row in df_prompts.iterrows():
            nome = row.get('Nome', row.get('nome', ''))
            tipo = row.get('Tipo', row.get('tipo', ''))
            status = row.get('Status', row.get('status', ''))
            instrucao = row.get('Instrução', row.get('Instrucao', row.get('instrucao', '')))
            row_id = row.get('ID', row.get('id', ''))
            
            c1, c2, c3, c4 = st.columns([3, 2, 1, 2])
            c1.markdown(f"**{nome}**")
            c2.markdown(f"`{tipo}`")
            
            if status == 'Ativo':
                c3.markdown("🟢 Ativo")
            else:
                c3.markdown("🔴 Inativo")
                
            with c4:
                col_btn_edit, col_btn_del = st.columns(2)
                with col_btn_edit:
                    edit_click = st.button("✏️ Ver/Editar", key=f"edit_prompt_{row_id}", use_container_width=True)
                with col_btn_del:
                    delete_click = st.button("🗑️ Excluir", key=f"del_prompt_{row_id}", use_container_width=True)
                    
            if delete_click:
                if db.excluir_prompt(row_id):
                    st.success("Prompt excluído com sucesso!")
                    st.session_state[f"editando_prompt_{row_id}"] = False
                    st.rerun()
                else:
                    st.error("Erro ao excluir prompt.")
            
            if edit_click:
                st.session_state[f"editando_prompt_{row_id}"] = not st.session_state.get(f"editando_prompt_{row_id}", False)
                
            if st.session_state.get(f"editando_prompt_{row_id}", False):
                with st.container():
                    with st.form(f"form_edit_prompt_{row_id}"):
                        st.write(f"**Editando Prompt:** {nome}")
                        col_ep1, col_ep2 = st.columns([2, 1])
                        with col_ep1:
                            edit_nome = st.text_input("Nome", value=nome, key=f"edit_nome_prompt_{row_id}")
                        with col_ep2:
                            edit_status_prompt = st.selectbox("Status", ["Ativo", "Inativo"], index=0 if status == 'Ativo' else 1, key=f"edit_status_prompt_{row_id}")
                            
                        edit_instrucao = st.text_area("Instrução", value=instrucao, height=250, key=f"edit_instrucao_prompt_{row_id}")
                        
                        c_btn1, c_btn2 = st.columns(2)
                        with c_btn1:
                            btn_salvar = st.form_submit_button("💾 Salvar", use_container_width=True)
                        with c_btn2:
                            btn_cancelar = st.form_submit_button("❌ Cancelar", use_container_width=True)
                            
                        if btn_salvar:
                            if db.atualizar_prompt(row_id, edit_nome, tipo, edit_instrucao, edit_status_prompt):
                                st.success("Atualizado!")
                                st.session_state[f"editando_prompt_{row_id}"] = False
                                st.rerun()
                            else:
                                st.error("Erro ao atualizar.")
                                
                        if btn_cancelar:
                            st.session_state[f"editando_prompt_{row_id}"] = False
                            st.rerun()
                                
            st.markdown("<hr style='margin: 5px 0px 10px 0px; border-top: 1px dashed #eee;'>", unsafe_allow_html=True)

# =====================================================================
# 4. TELA: IMPORTAÇÃO DE ARQUIVO DE DADOS
# =====================================================================
elif menu == "⚡ Tratar Planilha (Injetar UPM)":
    st.title("⚡ Tratar Planilha e Injetar UPM")
    st.write("Envie sua planilha Excel de Ocorrências (contendo as colunas **Bairro** e **Municipio**). O robô irá higienizar os nomes (corrigindo erros e substituindo apelidos) e criará automaticamente uma nova coluna com a **UPM** correspondente.")
    st.info("🧠 **Inteligência Artificial Ativa:** Se a planilha possuir as colunas **Narrativa** e **Tipo Local 2**, o robô analisará o texto e preencherá automaticamente a coluna **Tipo Local 2**, preservando o Tipo Local original.")
    
    def reset_processar():
        st.session_state.processar_clicado = False
        st.session_state.interromper = False
        # Limpa estados de execuções anteriores na sessão
        for k in ['df_upload_proc', 'qtd_bairros_corrigidos', 'qtd_municipios_corrigidos', 'qtd_upms_injetadas', 'colunas_data_fato', 'linhas_classificadas', 'linhas_mantidas', 'erros_ia', 'lista_erros_ia', 'lista_sucesso_ia', 'logs_ordenados']:
            if k in st.session_state:
                del st.session_state[k]
        
    # Formulário para agrupar as configurações e ações de processamento, impedindo reruns ao interagir com o checkbox
    with st.form("form_tratamento_planilha", border=False):
        # Checkbox profissional para ativação opcional da IA (fixo e desmarcado por padrão, sempre visível)
        usar_ia = st.checkbox(
            "🧠 Executar Classificação Cognitiva de Tipo Local via IA", 
            value=False, 
            help="Utiliza Inteligência Artificial (Gemini) para categorizar a coluna 'Tipo Local 2' com base no texto da Narrativa. Se desmarcado, o processamento fará apenas a higienização de nomes e injeção de UPMs."
        )
        
        uploaded_file = st.file_uploader("Escolha um arquivo XLS ou XLSX para tratamento", type=["xls", "xlsx"])
        
        st.write("")
        col_btn1, col_btn2, col_btn3 = st.columns([2, 2, 4])
        with col_btn1:
            btn_iniciar = st.form_submit_button("▶️ Iniciar Processamento", use_container_width=True)
        with col_btn2:
            btn_cancelar = st.form_submit_button("⏹️ Interromper / Cancelar", use_container_width=True)

    if btn_iniciar:
        if uploaded_file is None:
            st.warning("⚠️ Por favor, escolha um arquivo para tratamento antes de iniciar o processamento.")
        else:
            reset_processar()
            st.session_state.processar_clicado = True
            # st.rerun() não é chamado aqui para permitir o processamento imediato do arquivo com a requisição do form
            
    if btn_cancelar:
        st.session_state.interromper = True
        st.session_state.processar_clicado = False
        st.rerun()
            
    # Placeholder para resultados (prévia de dados e botão de download)
    placeholder_resultados = st.empty()
    if st.session_state.get('processar_clicado', False):
        placeholder_resultados.empty()

    # Permite continuar o fluxo se já possui dados processados na sessão (completo ou parcial interrompido)
    pode_exibir_resultado = 'df_upload_proc' in st.session_state
    if not st.session_state.get('processar_clicado', False) and not pode_exibir_resultado:
        if st.session_state.get('interromper', False):
            st.warning("⚠️ Processamento interrompido / cancelado com sucesso!")
        else:
            st.info("👆 Clique no botão acima para iniciar o tratamento da planilha.")
        st.stop()

    try:
        # 1. Inspeciona os tipos de célula da segunda linha (primeira linha de dados) usando openpyxl
        import openpyxl
        import re
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active

        linhas = list(ws.iter_rows(min_row=1, max_row=2))
        wb.close()

        dtype_dict = {}
        colunas_data = []  # Colunas que o Excel armazena como data

        # Padrões de number_format que indicam que a célula é uma data
        padrao_data = re.compile(r'[dDmMyY].*[dDmMyY]')

        if len(linhas) >= 2:
            header_row = linhas[0]  # Linha 1: cabeçalhos
            data_row = linhas[1]    # Linha 2: primeira linha de dados

            for header_cell, data_cell in zip(header_row, data_row):
                col_name = header_cell.value
                if col_name is None:
                    continue
                # Se a célula de dados é do tipo texto ('s') no Excel, forçar leitura como str
                if data_cell.data_type == 's':
                    dtype_dict[col_name] = str
                # Se a célula de dados é do tipo data ('d'), ou is_date, ou tem number_format de data
                elif data_cell.data_type == 'd' or data_cell.is_date:
                    colunas_data.append(col_name)
                elif data_cell.data_type == 'n' and data_cell.number_format and padrao_data.search(str(data_cell.number_format)):
                    colunas_data.append(col_name)

        # 2. Carrega os dados reais aplicando os dtypes detectados da planilha
        uploaded_file.seek(0)
        df_upload = pd.read_excel(uploaded_file, dtype=dtype_dict)

        # 3. Converte colunas de data que ficaram como número serial do Excel para datetime real
        #    Inclui as detectadas pelo openpyxl + fallback para qualquer coluna numérica inteira
        #    no range de datas seriais do Excel (cobre cenários que o openpyxl não detectou)
        for col in df_upload.columns:
            eh_coluna_data = col in colunas_data

            # Fallback: se a coluna não foi detectada pelo openpyxl mas contém apenas inteiros
            # no range de datas seriais do Excel (20000-80000 = anos 1954-2118), tratar como data
            if not eh_coluna_data and df_upload[col].dtype in ['int64', 'float64']:
                valores_validos = df_upload[col].dropna()
                if not valores_validos.empty:
                    todos_inteiros = all(float(v).is_integer() for v in valores_validos)
                    if todos_inteiros:
                        min_val = valores_validos.min()
                        max_val = valores_validos.max()
                        if 20000 < min_val and max_val < 80000:
                            eh_coluna_data = True
                            colunas_data.append(col)

            if eh_coluna_data:
                def converter_data_serial(val):
                    if pd.isna(val):
                        return val
                    # Se já for datetime, manter como está
                    if isinstance(val, (pd.Timestamp, datetime)):
                        return val
                    # Se for número serial do Excel (int ou float), converter para datetime real
                    try:
                        n = float(val)
                        if 1 < n < 100000:
                            return pd.to_datetime(n - 2, unit='D', origin='1900-01-01')
                    except (ValueError, TypeError):
                        pass
                    # Se for string no formato DD/MM/AAAA, converter para datetime
                    try:
                        return pd.to_datetime(val, dayfirst=True)
                    except Exception:
                        pass
                    return val
                df_upload[col] = df_upload[col].apply(converter_data_serial)

        # Guarda a lista de colunas de data para formatar na exportação
        st.session_state['colunas_data_detectadas'] = colunas_data

        # Verifica se a primeira coluna é RISP — se for, remove (será substituída por UPM)
        primeira_col = df_upload.columns[0]
        if str(primeira_col).strip().upper() == "RISP":
            df_upload = df_upload.drop(columns=[primeira_col])

        # Encontra colunas Bairro e Municipio de forma flexível (ignorando acentos e case)
        col_bairro = None
        col_municipio = None

        for col in df_upload.columns:
            col_norm = db.normalizar_texto(col)
            if col_norm == "bairro":
                col_bairro = col
            elif col_norm == "municipio":
                col_municipio = col

        if not col_bairro or not col_municipio:
            st.error("⚠️ Erro: A planilha enviada deve conter colunas chamadas **Bairro** e **Municipio** (ou Município).")
            st.info("Colunas encontradas no arquivo: " + ", ".join(df_upload.columns.tolist()))
        else:
            # Remove linhas onde bairro ou municipio estão vazios
            df_upload = df_upload.dropna(subset=[col_bairro, col_municipio])

            if df_upload.empty:
                st.warning("⚠️ Atenção: A planilha enviada não contém linhas de dados válidas (bairro e município vazios).")
            else:
                # Inicializa variáveis locais de estatísticas de IA com valor padrão para evitar NameError
                linhas_classificadas = 0
                linhas_mantidas = 0
                erros_ia = 0
                lista_erros_ia = []
                lista_sucesso_ia = []
                logs_ordenados = []

                if st.session_state.get('processar_clicado', False):
                    # Obtém o cache de UPMs na memória
                    # Obtém os caches na memória do banco de dados
                    mapa_upms = db.obter_mapeamento_upms()
                    mapa_nomes_mun = db.obter_mapeamento_nomes_municipios()
                    mapa_nomes_bai = db.obter_mapeamento_nomes_bairros()
                    mapa_alternativos_bai = db.obter_mapeamento_alternativo_bairros()
                    mapa_mun_todos = db.obter_municipios_com_bairro_todos_unico()

                    lista_upms = []
                    novos_bairros = []
                    novos_municipios = []

                    qtd_bairros_corrigidos = 0
                    qtd_municipios_corrigidos = 0
                    qtd_upms_injetadas = 0

                    for idx, row in df_upload.iterrows():
                        b_orig = row[col_bairro]
                        m_orig = row[col_municipio]

                        # 1. Padroniza e corrige o Município conforme as regras do usuário (TRIM + UPPER + apelidos)
                        m_padrao = db.padronizar_municipio(m_orig)
                        novos_municipios.append(m_padrao)

                        # 2. Padroniza e corrige o Bairro (TRIM + UPPER + mojibake)
                        b_padrao = db.corrigir_mojibake(b_orig).strip().upper()

                        # 3. Normaliza para fazer a busca no banco de dados
                        b_norm = db.normalizar_texto(b_padrao)
                        m_norm = db.normalizar_texto(m_padrao)

                        # 4. Mapeamento de equivalências para obter o Bairro Oficial (sempre em UPPER)
                        if (b_norm, m_norm) in mapa_nomes_bai:
                            b_oficial = mapa_nomes_bai[(b_norm, m_norm)].upper().strip()
                        elif (b_norm, m_norm) in mapa_alternativos_bai:
                            b_oficial = mapa_alternativos_bai[(b_norm, m_norm)].upper().strip()
                            # Atualiza a chave normalizada de busca do bairro para obter a UPM correta
                            b_norm = db.normalizar_texto(b_oficial)
                        else:
                            b_oficial = b_padrao

                        novos_bairros.append(b_oficial)

                        # 5. Determina a UPM correspondente usando a chave de bairro (oficial ou equivalente)
                        upm_val = mapa_upms.get((b_norm, m_norm))
                        if not upm_val:
                            # Se não achou Bairro+Municipio, mas o Município possui apenas 'TODOS' cadastrado
                            if m_norm in mapa_mun_todos:
                                upm_val = mapa_mun_todos[m_norm]
                            else:
                                upm_val = "NI"

                        if str(b_orig).strip().upper() != b_oficial:
                            qtd_bairros_corrigidos += 1
                        if str(m_orig).strip().upper() != m_padrao:
                            qtd_municipios_corrigidos += 1
                        if upm_val != "NI":
                            qtd_upms_injetadas += 1

                        lista_upms.append(upm_val)

                    # Atualiza o DataFrame com as strings higienizadas e corretas
                    df_upload[col_bairro] = novos_bairros
                    df_upload[col_municipio] = novos_municipios

                    # Cria a nova coluna UPM na primeira posição para ser facilmente visível
                    if "UPM" in df_upload.columns:
                        df_upload["UPM"] = lista_upms
                    else:
                        df_upload.insert(0, "UPM", lista_upms)

                    # =========================================================
                    # PASSO EXTRA: CLASSIFICAÇÃO INTELIGENTE (IA) TIPO LOCAL 2
                    # =========================================================
                    col_narrativa = next((c for c in df_upload.columns if db.normalizar_texto(c) == "narrativa"), None)
                    col_tipo2 = next((c for c in df_upload.columns if db.normalizar_texto(c).replace(" ", "") == "tipolocal2"), None)

                    if col_tipo2:
                        df_upload[col_tipo2] = df_upload[col_tipo2].astype(object)

                    st.session_state.df_upload_proc = df_upload
                    st.session_state.qtd_bairros_corrigidos = qtd_bairros_corrigidos
                    st.session_state.qtd_municipios_corrigidos = qtd_municipios_corrigidos
                    st.session_state.qtd_upms_injetadas = qtd_upms_injetadas
                    st.session_state.colunas_data_fato = colunas_data
                    st.session_state.linhas_classificadas = 0
                    st.session_state.linhas_mantidas = 0
                    st.session_state.erros_ia = 0
                    st.session_state.lista_erros_ia = []
                    st.session_state.lista_sucesso_ia = []
                    st.session_state.logs_ordenados = []
                else:
                    # Recupera os dados parciais da sessão
                    df_upload = st.session_state.df_upload_proc
                    qtd_bairros_corrigidos = st.session_state.get('qtd_bairros_corrigidos', 0)
                    qtd_municipios_corrigidos = st.session_state.get('qtd_municipios_corrigidos', 0)
                    qtd_upms_injetadas = st.session_state.get('qtd_upms_injetadas', 0)
                    colunas_data = st.session_state.get('colunas_data_fato', [])

                    linhas_classificadas = st.session_state.get('linhas_classificadas', 0)
                    linhas_mantidas = st.session_state.get('linhas_mantidas', 0)
                    erros_ia = st.session_state.get('erros_ia', 0)
                    lista_erros_ia = st.session_state.get('lista_erros_ia', [])
                    logs_ordenados = st.session_state.get('logs_ordenados', [])
                    lista_sucesso_ia = st.session_state.get('lista_sucesso_ia', [])

                    # Detecta as colunas de narrativa e tipo local a partir do DataFrame recuperado
                    col_narrativa = next((c for c in df_upload.columns if db.normalizar_texto(c) == "narrativa"), None) if df_upload is not None else None
                    col_tipo2 = next((c for c in df_upload.columns if db.normalizar_texto(c).replace(" ", "") == "tipolocal2"), None) if df_upload is not None else None
                if st.session_state.get('processar_clicado', False) and usar_ia and col_narrativa and col_tipo2 and ia.GENAI_DISPONIVEL and ia.GCP_PROJECT_ID and db.obter_prompt_ativo("Tipo Local"):
                    st.info("🧠 Inicializando Classificação Cognitiva de Tipo Local via IA...")
                    tipos_validos = db.obter_set_tipos_local()
                    linhas_classificadas = 0
                    linhas_mantidas = 0
                    erros_ia = 0
                    lista_erros_ia = []
                    lista_sucesso_ia = []

                    # Filtra em memória as ocorrências que realmente precisam de classificação por IA
                    indices_para_classificar = []
                    for row_idx, row_data in df_upload.iterrows():
                        valor_atual = str(row_data[col_tipo2]).strip().upper() if pd.notna(row_data[col_tipo2]) else ""
                        if not valor_atual or valor_atual not in tipos_validos:
                            narrativa = row_data[col_narrativa]
                            if pd.notna(narrativa) and str(narrativa).strip():
                                indices_para_classificar.append((row_idx, str(narrativa), valor_atual))
                            else:
                                linhas_mantidas += 1
                        else:
                                linhas_mantidas += 1

                    total_tarefas = len(indices_para_classificar)

                    if total_tarefas > 0:
                        import concurrent.futures
                        # 15 workers concorrentes para alta performance aproveitando as cotas do GCP
                        max_workers = 15

                        status_text = st.empty()
                        progress_bar = st.progress(0)
                        log_container = st.empty()

                        # Lista ordenada que guardará o log na posição exata da ocorrência (1-based index)
                        logs_ordenados = [None] * total_tarefas

                        with log_container.container():
                            with st.expander("📊 Log ao vivo das classificações da IA (Progresso Cronológico)", expanded=True):
                                st.info("⏳ Inicializando chamadas paralelas à Inteligência Artificial...")

                        # Função auxiliar executada de forma assíncrona nas threads filhas
                        def processar_linha_ia(row_idx, narrativa, valor_atual, num_tarefa):
                            texto_seguro = ia.anonimizar_texto(narrativa)
                            msg_log_inicio = f"🔒 Narrativa Anonimizada (Ocorrência {num_tarefa}): {texto_seguro}"

                            def callback_silencioso(msg):
                                pass

                            novo_tipo = ia.classificar("Tipo Local", narrativa, log_callback=callback_silencioso)
                            return row_idx, novo_tipo, valor_atual, msg_log_inicio, num_tarefa

                        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
                            # Submete as tarefas ao pool
                            futuros = {
                                executor.submit(processar_linha_ia, r_idx, narr, val_at, i + 1): r_idx
                                for i, (r_idx, narr, val_at) in enumerate(indices_para_classificar)
                            }

                            finalizados = 0
                            for futuro in concurrent.futures.as_completed(futuros):
                                # Se o usuário clicou para interromper, pára o processamento na thread principal
                                if st.session_state.get('interromper', False):
                                    st.warning("⚠️ Processamento de Inteligência Artificial interrompido pelo usuário!")
                                    executor.shutdown(wait=False, cancel_futures=True)
                                    break

                                try:
                                    row_idx, novo_tipo, valor_atual, msg_log_inicio, num_tarefa = futuro.result()

                                    msg_final = ""
                                    if novo_tipo.startswith("ERRO_"):
                                        erros_ia += 1
                                        lista_erros_ia.append(f"Ocorrência {num_tarefa}: {novo_tipo}")
                                        msg_final = f"🔒 Ocorrência {num_tarefa} ({ia.anonimizar_texto(df_upload.at[row_idx, col_narrativa])[:50]}...): Erro na análise ({novo_tipo})"
                                    elif novo_tipo == "NI":
                                        linhas_mantidas += 1
                                        msg_final = f"🔒 Ocorrência {num_tarefa}: Mantido [ {valor_atual if valor_atual else 'VAZIO'} ] -> A IA retornou NI (Não Identificado)"
                                    else:
                                        df_upload.at[row_idx, col_tipo2] = novo_tipo
                                        linhas_classificadas += 1
                                        msg_final = f"🔒 Ocorrência {num_tarefa}: Reclassificado de [ {valor_atual if valor_atual else 'VAZIO'} ] para -> [ {novo_tipo} ]"

                                    # Armazena na posição correspondente para o log final ordenado
                                    logs_ordenados[num_tarefa - 1] = msg_final

                                    # Coloca na lista cronológica de log ao vivo para manter o fluxo coerente sem pular números
                                    lista_sucesso_ia.append(msg_final)

                                    # Sincroniza em tempo real com o session_state para salvar progresso parcial em caso de cancelamento
                                    st.session_state.df_upload_proc = df_upload
                                    st.session_state.linhas_classificadas = linhas_classificadas
                                    st.session_state.linhas_mantidas = linhas_mantidas
                                    st.session_state.erros_ia = erros_ia
                                    st.session_state.lista_erros_ia = lista_erros_ia
                                    st.session_state.lista_sucesso_ia = lista_sucesso_ia
                                    st.session_state.logs_ordenados = logs_ordenados

                                except Exception as e:
                                    erros_ia += 1
                                    lista_erros_ia.append(f"Erro ao processar ocorrência: {str(e)}")

                                finalizados += 1
                                progress_bar.progress(finalizados / total_tarefas)
                                status_text.text(f"Processadas {finalizados} de {total_tarefas} ocorrências...")

                                # Exibe os logs ao vivo em ordem cronológica de finalização das requisições (sem pular ou sumir)
                                with log_container.container():
                                    with st.expander("📊 Log ao vivo das classificações da IA (Progresso Cronológico)", expanded=True):
                                        for msg in lista_sucesso_ia[-10:]: # Mostra os últimos 10 processados cronologicamente
                                            if "Reclassificado" in msg:
                                                st.success(msg)
                                            elif "Erro" in msg or "ERRO" in msg:
                                                st.error(msg)
                                            else:
                                                st.warning(msg)

                        status_text.empty()
                        progress_bar.empty()

                        # Renderiza o log final completo ordenado e fechado
                        logs_validos_finais = [msg for msg in logs_ordenados if msg is not None]
                        if logs_validos_finais:
                            with log_container.container():
                                with st.expander("📊 Ver log detalhado das classificações da IA (Ordenado)", expanded=False):
                                    for msg in logs_validos_finais:
                                        if "Reclassificado" in msg:
                                            st.success(msg)
                                        elif "Erro" in msg or "ERRO" in msg:
                                            st.error(msg)
                                        else:
                                            st.warning(msg)

                        if lista_erros_ia:
                            with st.expander("🚨 Ver detalhes dos erros da IA"):
                                for erro_msg in lista_erros_ia:
                                    st.error(erro_msg)

                elif col_narrativa and col_tipo2:
                    motivos_falta = []
                    if not ia.GENAI_DISPONIVEL:
                        motivos_falta.append("Pacote do Gemini não instalado")
                    if not ia.GCP_PROJECT_ID:
                        motivos_falta.append("ID do Projeto GCP não configurado no arquivo .env")
                    if not db.obter_prompt_ativo("Tipo Local"):
                        motivos_falta.append("Prompt para 'Tipo Local' não cadastrado ou inativo no sistema")

                    st.warning(f"⚠️ Colunas para IA encontradas, mas a classificação foi pulada. Motivos: {', '.join(motivos_falta)}.")

                # Concluiu o processamento, desativa o flag para evitar re-execuções reativas indesejadas
                st.session_state.processar_clicado = False

                # Envelopa toda a exibição dos resultados e prévia no container vazio do placeholder_resultados
                with placeholder_resultados.container():
                    # Mensagem de conclusão compacta e organizada
                    if st.session_state.get('interromper', False):
                        st.warning("⚠️ Processamento Parcial: IA interrompida pelo usuário. Abaixo estão disponíveis os dados analisados até a interrupção.")
                    else:
                        st.success("🎉 Processamento Concluído com Sucesso!")
                    qtd_upms_ni = len(df_upload) - qtd_upms_injetadas

                    st.info(f"🎯 **Tratamento (Bairros e Municípios):** {len(df_upload)} linhas | {qtd_municipios_corrigidos} Municípios Higienizados | {qtd_bairros_corrigidos} Bairros Corrigidos | {qtd_upms_injetadas} UPMs Injetadas | {qtd_upms_ni} UPMs 'NI'")

                    if col_narrativa and col_tipo2 and ia.GENAI_DISPONIVEL:
                        st.info(f"🧠 **Classificação por Inteligência Artificial:** {linhas_classificadas} Reclassificações | {linhas_mantidas} Mantidos (Sem Alteração) | {erros_ia} Erros")

                    st.subheader("Prévia dos Dados Processados")
                    # Exibe no Streamlit apenas as primeiras 100 linhas como prévia para performance
                    linhas_prev = min(100, len(df_upload))
                    st.write(f"Mostrando as primeiras {linhas_prev} de {len(df_upload)} linhas processadas:")

                    df_preview = df_upload.head(linhas_prev).copy()
                    df_preview.index = range(1, len(df_preview) + 1)

                    # Formata colunas de data como DD/MM/AAAA para exibição na grid
                    for col_dt in colunas_data:
                        if col_dt in df_preview.columns:
                            df_preview[col_dt] = df_preview[col_dt].apply(
                                lambda v: v.strftime('%d/%m/%Y') if isinstance(v, (pd.Timestamp, datetime)) else v
                            )

                    st.dataframe(df_preview, width="stretch")

                    # Cria o arquivo Excel em memória para download
                    import io
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl', datetime_format='dd/mm/yyyy') as writer:
                        df_upload.to_excel(writer, index=False, sheet_name='Dados_Processados')

                        # Aplica formato dd/mm/yyyy nas colunas de data detectadas
                        colunas_data_exp = st.session_state.get('colunas_data_detectadas', [])
                        if colunas_data_exp:
                            ws = writer.sheets['Dados_Processados']
                            from openpyxl.utils import get_column_letter
                            colunas_df = list(df_upload.columns)
                            for col_data_nome in colunas_data_exp:
                                if col_data_nome in colunas_df:
                                    col_idx = colunas_df.index(col_data_nome) + 1  # +1 pois Excel é 1-indexed
                                    col_letter = get_column_letter(col_idx)
                                    for row in range(2, len(df_upload) + 2):  # +2: header na linha 1
                                        cell = ws[f'{col_letter}{row}']
                                        cell.number_format = 'dd/mm/yyyy'
                    # Determina o nome do arquivo dinâmico (ROUBO E FURTO data_min A data_max - EXPORT.xlsx)
                    nome_arquivo = f"ROUBO E FURTO {datetime.today().strftime('%d-%m-%Y')} - EXPORT.xlsx"
                    col_data_fato = None
                    for col in df_upload.columns:
                        # Remove espaços e underlines para a comparação (Ex: "Data Fato" -> "datafato")
                        col_comparar = db.normalizar_texto(col).replace(" ", "").replace("_", "")
                        if col_comparar == "datafato":
                            col_data_fato = col
                            break

                    if col_data_fato:
                        # Converte para datetime garantindo o padrão brasileiro (dia primeiro)
                        datas_validas = pd.to_datetime(df_upload[col_data_fato], dayfirst=True, errors='coerce').dropna()
                        if not datas_validas.empty:
                            data_min = datas_validas.min().strftime('%d-%m-%Y')
                            data_max = datas_validas.max().strftime('%d-%m-%Y')
                            nome_arquivo = f"ROUBO E FURTO {data_min} A {data_max} - EXPORT.xlsx"

                    processed_data = buffer.getvalue()

                    st.write("")
                    st.download_button(
                        label="📥 Baixar Planilha com Coluna UPM (.xlsx)",
                        data=processed_data,
                        file_name=nome_arquivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width="stretch"
                    )

    except Exception as e:
        st.error(f"⚠️ Erro ao ler ou processar o arquivo: {str(e)}")
    else:
        st.info("💡 Dica: Prepare uma planilha contendo as colunas 'Bairro' e 'Municipio' para cruzar com o banco de dados do sistema.")

# =====================================================================
# 4.5. TELA: EXTRAÇÃO DE PDFs (BO)
# =====================================================================
elif menu == "🤖 Robô de Extração (BO)":
    st.title("📥 Extração de Dados de BOs (PDF)")
    st.write("Faça o upload de um ou mais arquivos PDF de Boletins de Ocorrência para extrair dados automaticamente por seções e exportar o resultado para Excel.")

    # Dropdown de Layout
    df_layouts = db.listar_layouts()
    lista_layouts = df_layouts["Nome_Layout"].tolist() if not df_layouts.empty else ["Layout Genérico Padrão"]
    layout_selecionado = st.selectbox("Selecione o Layout de Extração", lista_layouts, key="combo_layout_manual")
    layout_id = int(df_layouts[df_layouts["Nome_Layout"] == layout_selecionado]["ID"].values[0]) if not df_layouts.empty else 1

    uploaded_files = st.file_uploader("Escolha os arquivos PDF dos BOs", type=["pdf"], accept_multiple_files=True, key="bo_pdf_uploader")

    if uploaded_files:
        if st.button("🚀 Processar BOs Selecionados", type="primary", width="stretch", key="btn_processar_bo_pdfs"):
            try:
                import pypdf
                import extrair_bo as ex_bo
                import importlib
                importlib.reload(ex_bo)
                
                # Inicializa ou garante que o banco está pronto
                db.inicializar_banco()
                
                # Carrega os mapeamentos e listas de bairros/municípios
                db_mappings = {
                    "mapa_upms": db.obter_mapeamento_upms(),
                    "mapa_nomes_mun": db.obter_mapeamento_nomes_municipios(),
                    "mapa_nomes_bai": db.obter_mapeamento_nomes_bairros(),
                    "mapa_alternativos_bai": db.obter_mapeamento_alternativo_bairros(),
                    "mapa_mun_todos": db.obter_municipios_com_bairro_todos_unico()
                }
                
                lista_municipios, bairros_por_mun = ex_bo.carregar_dados_banco()
                
                resultados = []
                progresso = st.progress(0)
                status_text = st.empty()
                
                total_arquivos = len(uploaded_files)
                for idx, uploaded_file in enumerate(uploaded_files):
                    status_text.write(f"Processando arquivo {idx+1} de {total_arquivos}: **{uploaded_file.name}**...")
                    
                    texto = ""
                    try:
                        texto = ex_bo.extrair_texto_pdf(uploaded_file)
                    except Exception as e:
                        texto = ""
                        st.error(f"Erro ao ler {uploaded_file.name}: {str(e)}")
                        
                    if not texto:
                        res = {
                            "ARQUIVO": uploaded_file.name,
                            "BO_NUMERO": "Erro na leitura",
                            "NARRATIVA": "Erro ao extrair texto do arquivo PDF.",
                            "PROVIDENCIAS": "NI"
                        }
                    else:
                        # Processa usando a lógica unificada do script extrair_bo
                        res = ex_bo.processar_texto_bo(
                            texto, 
                            uploaded_file.name,
                            layout_id,
                            db_mappings, 
                            lista_municipios, 
                            bairros_por_mun
                        )
                    resultados.append(res)
                    progresso.progress((idx + 1) / total_arquivos)
                
                status_text.empty()
                st.success(f"🎉 Extração concluída! {len(resultados)} arquivos processados com sucesso.")
                
                df_res = pd.DataFrame(resultados)
                # Ordena as colunas usando a regra unificada centralizada no script
                df_res = ex_bo.ordenar_dataframe(df_res, layout_id)
                
                # Ajusta o índice para iniciar em 1 para exibição amigável
                df_res.index = df_res.index + 1
                
                st.subheader("Prévia dos BOs Extraídos")
                st.dataframe(df_res, width="stretch")
                
                # Gera o buffer do Excel para download
                import io
                buffer_pdf_xlsx = io.BytesIO()
                with pd.ExcelWriter(buffer_pdf_xlsx, engine='openpyxl') as writer:
                    df_res.to_excel(writer, index=False, sheet_name='BOs_Extraidos')
                    
                nome_xlsx = f"BOs_Processados_{datetime.today().strftime('%d-%m-%Y')}.xlsx"
                
                st.download_button(
                    label="📥 Baixar Planilha de BOs (.xlsx)",
                    data=buffer_pdf_xlsx.getvalue(),
                    file_name=nome_xlsx,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    width="stretch",
                    key="btn_download_bo_xlsx"
                )
            except Exception as e:
                st.error(f"Erro ao processar os arquivos PDF: {str(e)}")

# =====================================================================
# 5. TELA: CARGA E CONFIGURAÇÕES
# =====================================================================
elif menu == "⚙️ Configurações":
    st.title("⚙️ Carga e Configurações")
    
    st.write("Gerencie a base de dados do sistema, baixe o arquivo de modelo para preenchimento, realize cargas em lote ou limpe o banco de dados completamente.")
    
    tab_carga, tab_export, tab_limpeza = st.tabs(["📥 Carga Geral de Dados", "📤 Exportar Banco", "⚠️ Redefinir Banco"])
    
    with tab_carga:
        st.subheader("1. Baixar Arquivo de Modelo")
        st.write("Baixe a planilha modelo contendo a estrutura correta para importar seus dados. Preencha as abas conforme necessário.")
        
        # Cria o modelo Excel em memória para download
        import io
        buffer_model = io.BytesIO()
        with pd.ExcelWriter(buffer_model, engine='openpyxl') as writer:
            pd.DataFrame(columns=["Municipio", "Estado"]).to_excel(writer, index=False, sheet_name="Municipios")
            pd.DataFrame(columns=["Bairro", "Municipio"]).to_excel(writer, index=False, sheet_name="Bairros")
            pd.DataFrame(columns=["Bairro_Oficial", "Municipio", "Nome_Alternativo"]).to_excel(writer, index=False, sheet_name="Nomes Alternativos")
            pd.DataFrame(columns=["UPM", "Descricao", "Bairro", "Municipio", "Estado"]).to_excel(writer, index=False, sheet_name="UPMs")
            pd.DataFrame(columns=["Nome", "UrlLogin", "UrlConsulta", "UrlPdf", "Login", "Senha", "DuplaAutenticacao", "Tipo", "Status"]).to_excel(writer, index=False, sheet_name="Servicos")
        processed_model = buffer_model.getvalue()
        
        st.download_button(
            label="📥 Baixar Modelo de Importação (.xlsx)",
            data=processed_model,
            file_name="modelo_carga_buscadados.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.markdown("---")
        
        st.subheader("2. Enviar Arquivo Preenchido")
        arquivo_carga = st.file_uploader("Escolha o arquivo Excel (.xlsx) preenchido", type=["xlsx"], key="file_carga_geral")
        
        if arquivo_carga is not None:
            if st.button("🚀 Processar Carga em Lote", type="primary", width="stretch"):
                try:
                    with st.spinner("Processando importação..."):
                        xls = pd.ExcelFile(arquivo_carga)
                        abas = xls.sheet_names
                        
                        # Validação inicial de abas
                        required_sheets = ["Municipios", "Bairros", "Nomes Alternativos", "UPMs", "Servicos"]
                        missing = [s for s in required_sheets if s not in abas]
                        
                        if missing:
                            st.error(f"Erro: O arquivo de carga deve conter as abas: {', '.join(required_sheets)}. Abas ausentes no arquivo: {', '.join(missing)}")
                        else:
                            # Carrega os DataFrames
                            df_mun = pd.read_excel(xls, sheet_name="Municipios") if "Municipios" in xls.sheet_names else pd.DataFrame()
                            df_bai = pd.read_excel(xls, sheet_name="Bairros") if "Bairros" in xls.sheet_names else pd.DataFrame()
                            df_alt = pd.read_excel(xls, sheet_name="Nomes Alternativos") if "Nomes Alternativos" in xls.sheet_names else pd.DataFrame()
                            df_upm = pd.read_excel(xls, sheet_name="UPMs") if "UPMs" in xls.sheet_names else pd.DataFrame()
                            df_serv = pd.read_excel(xls, sheet_name="Servicos") if "Servicos" in xls.sheet_names else pd.DataFrame()
                            df_layouts = pd.read_excel(xls, sheet_name="Layouts") if "Layouts" in xls.sheet_names else pd.DataFrame()
                            df_grupos = pd.read_excel(xls, sheet_name="Grupos") if "Grupos" in xls.sheet_names else pd.DataFrame()
                            df_itens = pd.read_excel(xls, sheet_name="Itens") if "Itens" in xls.sheet_names else pd.DataFrame()
                            df_tipos_local = pd.read_excel(xls, sheet_name="Tipos Local") if "Tipos Local" in xls.sheet_names else pd.DataFrame()
                            df_prompts = pd.read_excel(xls, sheet_name="Prompts") if "Prompts" in xls.sheet_names else pd.DataFrame()
                            
                            # Executa as cargas na ordem correta
                            res_mun = db.importar_municipios_lote(df_mun)
                            res_bai = db.importar_bairros_lote(df_bai)
                            res_alt = db.importar_nomes_alternativos_lote(df_alt)
                            res_upm = db.importar_upms_lote(df_upm)
                            res_serv = db.importar_servicos_lote(df_serv)
                            res_lay = db.importar_layouts_lote(df_layouts)
                            res_grp = db.importar_grupos_lote(df_grupos)
                            res_itn = db.importar_itens_lote(df_itens)
                            
                            res_tipos_local = {"inseridos": 0, "pulados": 0, "erros": 0}
                            res_prompts = {"inseridos": 0, "pulados": 0, "erros": 0}
                            if not df_tipos_local.empty:
                                res_tipos_local = db.importar_tipos_local_lote(df_tipos_local)
                            if not df_prompts.empty:
                                res_prompts = db.importar_prompts_ia_lote(df_prompts)
                            
                            # Correção de integridade para PostgreSQL
                            db.sincronizar_sequencias()
                            
                            st.success("🎉 Carga de dados realizada com sucesso!")
                            
                            # Exibe o resultado de forma visualmente rica
                            st.markdown("### Resumo da Carga")
                            col1, col2, col3, col4, col5 = st.columns(5)
                            with col1:
                                st.metric("Municípios", f"+{res_mun['inseridos']}", f"Ignorados: {res_mun['pulados']} | Erros: {res_mun['erros']}")
                            with col2:
                                st.metric("Bairros", f"+{res_bai['inseridos']}", f"Ignorados: {res_bai['pulados']} | Erros: {res_bai['erros']}")
                            with col3:
                                st.metric("UPMs Mapeadas", f"+{res_upm['inseridos']}", f"Ignorados: {res_upm['pulados']} | Erros: {res_upm['erros']}")
                            with col4:
                                st.metric("Serviços", f"+{res_serv['inseridos']}", f"Ignorados: {res_serv['pulados']} | Erros: {res_serv['erros']}")
                            with col5:
                                total_ia = res_tipos_local.get('inseridos', 0) + res_prompts.get('inseridos', 0)
                                erros_ia = res_tipos_local.get('erros', 0) + res_prompts.get('erros', 0)
                                st.metric("Config. IA", f"+{total_ia}", f"Erros: {erros_ia}")
                except Exception as e:
                    st.error(f"⚠️ Erro ao processar arquivo: {str(e)}")
                    
    with tab_export:
        st.subheader("📤 Exportar Todos os Dados (Backup em Excel)")
        st.write("Baixe um arquivo Excel contendo todos os dados atualmente cadastrados no banco (Municípios, Bairros, Nomes Alternativos, UPMs e Serviços). Este arquivo possui exatamente a mesma estrutura aceita na aba de Importação e pode ser usado para clonar ou restaurar o sistema.")
        
        # Como a geração pode demorar, fazemos com um botão antes do download real
        if st.button("🔄 Compilar Dados para Exportação", type="secondary", use_container_width=True):
            with st.spinner("Extraindo e formatando dados do banco..."):
                try:
                    conn_exp = db.engine
                    
                    df_mun_exp = db.ajustar_colunas(pd.read_sql("SELECT Municipio, Estado, id_municipio_srop FROM municipios", conn_exp))
                    df_bai_exp = db.ajustar_colunas(pd.read_sql("""
                        SELECT b.Bairro, m.Municipio
                        FROM bairros b
                        JOIN municipios m ON b.Municipio_ID = m.ID
                        ORDER BY m.Municipio, b.Bairro
                    """, conn_exp))
                    df_alt_exp = db.ajustar_colunas(pd.read_sql("SELECT b.Bairro as Bairro_Oficial, m.Municipio, a.Nome_Alternativo FROM bairros_alternativos a JOIN bairros b ON a.Bairro_ID = b.ID JOIN municipios m ON b.Municipio_ID = m.ID", conn_exp))
                    df_upm_exp = pd.read_sql("""
                        SELECT 
                            u.UPM, 
                            u.Descricao,
                            b.Bairro,
                            m.Municipio,
                            m.Estado
                        FROM upms u 
                        JOIN upm_bairros ub ON u.ID = ub.UPM_ID 
                        JOIN bairros b ON ub.Bairro_ID = b.ID
                        JOIN municipios m ON b.Municipio_ID = m.ID
                    """, conn_exp)
                    df_serv_exp = db.ajustar_colunas(pd.read_sql("SELECT Nome, UrlLogin, UrlConsulta, UrlPdf, Login, Senha, DuplaAutenticacao, Tipo, Status, Exibir_No_Menu FROM servicos", conn_exp))
                    
                    df_layouts_exp = db.ajustar_colunas(pd.read_sql("SELECT Nome_Layout FROM layouts", conn_exp))
                    df_grupos_exp = db.ajustar_colunas(pd.read_sql("SELECT l.Nome_Layout, g.Nome_Grupo, g.Ordem, g.Ordem_Excel, g.Tem_Itens, g.Exportar_Excel FROM layout_grupos g JOIN layouts l ON g.Layout_ID = l.ID", conn_exp))
                    df_itens_exp = db.ajustar_colunas(pd.read_sql("SELECT l.Nome_Layout, g.Nome_Grupo, i.Nome_Item_Excel, i.Palavra_Busca, i.Ordem, i.Ordem_Excel, i.Exportar_Excel FROM layout_itens i JOIN layout_grupos g ON i.Grupo_ID = g.ID JOIN layouts l ON g.Layout_ID = l.ID", conn_exp))
                    
                    df_tipos_local_exp = db.ajustar_colunas(pd.read_sql("SELECT Tipo_Local, Descricao_IA, Status FROM tipos_local", conn_exp))
                    df_prompts_exp = db.ajustar_colunas(pd.read_sql("SELECT Nome, Tipo, Instrucao, Status FROM prompts_ia", conn_exp))
                    
                    # Não é necessário fechar a engine do SQLAlchemy
                    
                    import io
                    buffer_export = io.BytesIO()
                    with pd.ExcelWriter(buffer_export, engine='openpyxl') as writer:
                        df_mun_exp.to_excel(writer, index=False, sheet_name="Municipios")
                        df_bai_exp.to_excel(writer, index=False, sheet_name="Bairros")
                        df_alt_exp.to_excel(writer, index=False, sheet_name="Nomes Alternativos")
                        df_upm_exp.to_excel(writer, index=False, sheet_name="UPMs")
                        df_serv_exp.to_excel(writer, index=False, sheet_name="Servicos")
                        df_layouts_exp.to_excel(writer, index=False, sheet_name="Layouts")
                        df_grupos_exp.to_excel(writer, index=False, sheet_name="Grupos")
                        df_itens_exp.to_excel(writer, index=False, sheet_name="Itens")
                        df_tipos_local_exp.to_excel(writer, index=False, sheet_name="Tipos Local")
                        df_prompts_exp.to_excel(writer, index=False, sheet_name="Prompts")
                    
                    st.session_state.export_ready_data = buffer_export.getvalue()
                    st.success("✅ Banco compilado com sucesso! Clique no botão abaixo para baixar.")
                except Exception as e:
                    st.error(f"Erro ao compilar banco: {str(e)}")
                    
        if "export_ready_data" in st.session_state and st.session_state.export_ready_data:
            st.download_button(
                label="📦 Baixar Backup em Excel (.xlsx)",
                data=st.session_state.export_ready_data,
                file_name="backup_buscadados_completo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )

    with tab_limpeza:
        st.subheader("🚨 Excluir Todos os Dados do Sistema")
        st.error("ATENÇÃO: Esta operação é irreversível e irá deletar TODOS os municípios, bairros, nomes alternativos, UPMs e vínculos cadastrados no banco de dados!")
        
        confirmacao = st.checkbox("Estou ciente e desejo apagar PERMANENTEMENTE todos os dados cadastrados.", key="confirmacao_reset_db")
        
        if st.button("🔴 Apagar Todos os Dados", disabled=not confirmacao, type="primary", width="stretch"):
            try:
                db.limpar_banco_dados()
                st.success("🔥 Todos os dados foram excluídos com sucesso. O banco de dados foi reiniciado.")
                st.balloons()
            except Exception as e:
                st.error(f"Erro ao limpar banco de dados: {str(e)}")

# =====================================================================
# 6. TELA DE EXECUÇÃO DINÂMICA DE SERVIÇOS
# =====================================================================
elif menu.startswith("servico_"):
    # Extrai o ID do serviço
    id_servico = int(menu.replace("servico_", ""))
    
    # Busca detalhes do serviço no banco
    conn = db.obter_conexao()
    cursor = conn.cursor()
    cursor.execute("SELECT Nome, UrlLogin, UrlConsulta, UrlPdf, Login, Senha, DuplaAutenticacao, Tipo, Status FROM servicos WHERE ID = %s", (id_servico,))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        st.error("⚠️ Serviço não encontrado.")
    else:
        nome, url_login, url_consulta, url_pdf, login_db, senha_db_cripto, dupla_autenticacao, tipo, status_db = row
        
        if status_db == "Inativo":
            st.error("⚠️ Este serviço está inativo no momento. Ative-o na tela de Manutenção de Serviços para poder executá-lo.")
            if st.button("⬅️ Voltar para Manutenção de Serviços", key="btn_voltar_inativo", width="stretch"):
                st.session_state.menu_override = None
                st.session_state.radio_selecionado = "🔌 Manutenção de Serviços"
                st.rerun()
            st.stop()
        
        st.title(f"🔌 Execução de Serviço: {nome}")
        st.write(f"Tipo de Serviço: **{tipo}**")
        
        # Só desenvolvemos o tipo SROP por enquanto
        if tipo != "SROP":
            st.warning(f"A integração para o tipo de serviço '{tipo}' não está implementada ainda. Apenas o tipo 'SROP' é suportado atualmente.")
        else:
            # Descriptografa a senha para usar ou exibir
            senha_db = db.descriptografar_senha(senha_db_cripto)
            
            cookie_key = f"srop_cookies_{id_servico}"
            result_key = f"resultado_exec_{id_servico}"
            
            # Tenta recuperar sessão ativa do banco caso não esteja na memória do Streamlit
            if cookie_key not in st.session_state:
                sessao_banco = db.obter_sessao_ativa(id_servico)
                if sessao_banco:
                    st.session_state[cookie_key] = sessao_banco["cookies"]
                    st.session_state[f"srop_login_time_{id_servico}"] = sessao_banco["data_login"]
            
            # Se não estiver conectado (sem cookies salvos)
            if cookie_key not in st.session_state:
                st.markdown("### 🔑 Credenciais e Login")
                
                # Se Login e Senha estiverem preenchidos no banco de dados, não exibe na tela
                has_credentials = bool(login_db.strip() and senha_db.strip())
                
                if has_credentials:
                    st.info("💡 As credenciais de acesso (Usuário e Senha) já estão pré-configuradas no banco de dados para este serviço e serão utilizadas automaticamente.")
                    usuario_exec = login_db.strip()
                    senha_exec = senha_db.strip()
                else:
                    st.warning("⚠️ Este serviço não possui Usuário e Senha pré-configurados no banco. Por favor, insira-os abaixo para executar o login.")
                    col_user, col_pass = st.columns(2)
                    with col_user:
                        usuario_exec = st.text_input("Usuário / Login", value="", key=f"exec_user_{id_servico}").strip()
                    with col_pass:
                        senha_exec = st.text_input("Senha", value="", type="password", key=f"exec_pass_{id_servico}").strip()
                
                codigo_mfa = ""
                if dupla_autenticacao == "Sim":
                    codigo_mfa = st.text_input("Código do Autenticador (MFA)", value="", placeholder="Ex: 123456", key=f"exec_mfa_{id_servico}").strip()
                
                st.write("")
                
                if st.button("🔑 Realizar Login no Serviço", type="primary", width="stretch", key=f"btn_login_{id_servico}"):
                    if not usuario_exec or not senha_exec:
                        st.error("⚠️ Erro: Usuário e Senha são necessários para realizar o acesso ao serviço.")
                    elif dupla_autenticacao == "Sim" and not codigo_mfa:
                        st.error("⚠️ Erro: A dupla autenticação está ativada. Por favor, preencha o Código do Autenticador (MFA).")
                    else:
                        with st.status("Autenticando no serviço...", expanded=True) as status_ui:
                            def atualizar_status_login(mensagem: str):
                                status_ui.update(label=mensagem, state="running", expanded=True)
                                st.write(mensagem)
                            
                            try:
                                import automacao as aut
                                session_state_data = aut.realizar_login_srop(
                                    url_login=url_login,
                                    usuario=usuario_exec,
                                    senha=senha_exec,
                                    codigo_mfa=codigo_mfa,
                                    dupla_autenticacao=(dupla_autenticacao == "Sim"),
                                    status_callback=atualizar_status_login
                                )
                                
                                if session_state_data:
                                    st.session_state[cookie_key] = session_state_data
                                    # Salva a sessão no banco
                                    db.salvar_sessao(id_servico, session_state_data)
                                    # Grava a data no streamlit para exibir na tela imediatamente
                                    from datetime import datetime
                                    st.session_state[f"srop_login_time_{id_servico}"] = datetime.now()
                                    
                                    status_ui.update(label="Login realizado com sucesso!", state="complete", expanded=True)
                                    st.success("🟢 Sessão iniciada e persistida com sucesso!")
                                    st.rerun()
                                else:
                                    status_ui.update(label="Falha ao obter estado de sessão.", state="error", expanded=True)
                                    st.error("❌ Não foi possível obter o estado da sessão de login.")
                            except Exception as error_login:
                                status_ui.update(label="Ocorreu um erro ao realizar o login.", state="error", expanded=True)
                                st.error(f"❌ {str(error_login)}")
            else:
                # Se estiver conectado (cookies salvos)
                sessao_info = "🟢 Conectado ao Serviço SROP"
                if f"srop_login_time_{id_servico}" in st.session_state:
                    from datetime import datetime
                    dt_login = st.session_state[f"srop_login_time_{id_servico}"]
                    if isinstance(dt_login, str):
                        try:
                            dt_login = datetime.strptime(dt_login, '%Y-%m-%d %H:%M:%S')
                        except:
                            dt_login = datetime.now()
                    diff = datetime.now() - dt_login
                    total_seconds = int(diff.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    tempo_ativo_fmt = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                    logado_em_fmt = dt_login.strftime('%d/%m/%Y %H:%M:%S')
                    sessao_info = f"🟢 Conectado ao Serviço SROP (Logado em: {logado_em_fmt} | Tempo Ativo: {tempo_ativo_fmt} | Limite: 4 Horas)"
                
                st.success(sessao_info)
                
                st.markdown("### 📅 Parâmetros de Consulta")
                from datetime import timedelta
                ontem = datetime.today() - timedelta(days=1)
                
                col_data_ini, col_data_fim = st.columns(2)
                with col_data_ini:
                    data_inicial = st.date_input("Data Inicial do Registro", value=ontem, format="DD/MM/YYYY", key=f"exec_dt_ini_{id_servico}")
                with col_data_fim:
                    data_final = st.date_input("Data Final do Registro", value=ontem, format="DD/MM/YYYY", key=f"exec_dt_fim_{id_servico}")
                
                # Novos campos do SROP
                df_muns_sel = db.listar_dados("municipios")
                options_mun = []
                idx_default_mun = 0
                df_muns_filtrado = pd.DataFrame()
                
                if not df_muns_sel.empty:
                    # Filtra apenas os municípios que possuem id_municipio_srop preenchido
                    df_muns_filtrado = df_muns_sel[df_muns_sel["id_municipio_srop"].notna() & (df_muns_sel["id_municipio_srop"].str.strip() != "")]
                    
                    if not df_muns_filtrado.empty:
                        # Ordena os municípios por nome alfabeticamente
                        df_muns_filtrado = df_muns_filtrado.sort_values(by="Municipio")
                        
                        # Separa o município com ID SROP igual a 9167 para colocar no topo da lista
                        df_9167 = df_muns_filtrado[df_muns_filtrado["id_municipio_srop"].str.strip() == "9167"]
                        df_outros = df_muns_filtrado[df_muns_filtrado["id_municipio_srop"].str.strip() != "9167"]
                        
                        # Reconstrói o DataFrame com o 9167 no topo
                        df_muns_filtrado = pd.concat([df_9167, df_outros]).reset_index(drop=True)
                        
                        options_mun = [f"{row['Municipio']} - {row['Estado']}" for _, row in df_muns_filtrado.iterrows()]
                        # O índice padrão será 0 (o primeiro item da lista) que agora corresponde ao 9167 se ele existir
                        idx_default_mun = 0
                            
                st.markdown("<div style='margin-top: 15px;'></div>", unsafe_allow_html=True)
                col_mun_sel, col_size_sel, col_bo_sel = st.columns([2, 1, 1])
                
                with col_mun_sel:
                    if options_mun:
                        mun_selecionado_str = st.selectbox("Município (SROP)", options_mun, index=idx_default_mun, key=f"exec_mun_srop_{id_servico}")
                        idx_sel = options_mun.index(mun_selecionado_str)
                        id_municipio_val = df_muns_filtrado.iloc[idx_sel].get("id_municipio_srop", "")
                    else:
                        st.selectbox("Município (SROP)", ["Nenhum município com ID SROP cadastrado"], key=f"exec_mun_srop_empty_{id_servico}", disabled=True)
                        id_municipio_val = ""
                        
                with col_size_sel:
                    size_val = st.number_input("Total de Registros (Size)", min_value=1, max_value=1000, value=1, step=10, key=f"exec_size_{id_servico}")
                    
                with col_bo_sel:
                    st.text_input("Número do BO (Em Desenvolvimento)", value="", disabled=True, key=f"exec_bo_num_{id_servico}")
                    bo_val = None
                    
                st.checkbox("Filtrar por Data de Ocorrência Fato (Em Desenvolvimento)", value=False, disabled=True, key=f"exec_chk_fato_{id_servico}")
                data_ini_fato_str = None
                data_fim_fato_str = None
                
                st.markdown("<p style='font-size: 13px; color: #888888; font-style: italic; margin-top: -5px;'>ℹ️ Os filtros de 'Número do BO' e 'Data do Fato' estão inativos na interface e não serão enviados na consulta pois estão em fase de homologação da nova API SROP.</p>", unsafe_allow_html=True)
                st.write("")
                
                # Trava de Segurança para períodos longos
                from datetime import timedelta
                periodo_longo = (data_final - data_inicial) > timedelta(days=1)
                pode_executar = True
                
                if periodo_longo:
                    st.warning("⚠️ **Aviso de Sobrecarga:** O período selecionado é maior que 1 dia! Consultas longas podem retornar centenas de registros e demorar muito tempo para serem concluídas.")
                    if not st.checkbox("Estou ciente e desejo executar a consulta mesmo assim.", key=f"chk_ciente_{id_servico}"):
                        pode_executar = False
                
                col_btn_run, col_btn_download_only, col_btn_cancel, col_btn_logoff = st.columns(4)
                
                with col_btn_run:
                    btn_run = st.button("🚀 Consultar e Extrair Dados", type="primary", use_container_width=True, disabled=not pode_executar, key=f"btn_run_{id_servico}")
                with col_btn_download_only:
                    btn_download_only = st.button("📥 Consultar e Baixar PDFs", use_container_width=True, disabled=not pode_executar, key=f"btn_download_only_{id_servico}")
                with col_btn_cancel:
                    placeholder_cancel = st.empty()
                    btn_cancel = placeholder_cancel.button("⏹️ Cancelar Processamento", use_container_width=True, disabled=not (btn_run or btn_download_only), key=f"btn_interrupt_{id_servico}")
                with col_btn_logoff:
                    btn_logoff = st.button("🔴 Encerrar Sessão", use_container_width=True, disabled=(btn_run or btn_download_only), key=f"btn_logoff_{id_servico}")
                
                if btn_logoff:
                    with st.spinner("Enviando comando de logoff para o servidor..."):
                        import automacao as aut
                        aut.encerrar_sessao_srop(st.session_state[cookie_key], url_login)
                        db.limpar_sessao(id_servico)
                        st.session_state.pop(cookie_key, None)
                        st.session_state.pop(result_key, None)
                        st.session_state.pop(f"srop_login_time_{id_servico}", None)
                        st.session_state.pop(f"zip_{result_key}", None)
                    st.success("Sessão encerrada com sucesso no servidor e no banco!")
                    st.rerun()
                
                if btn_run:
                    # Validação em tempo de execução para garantir que as tags obrigatórias estão na URL base
                    tags_obrigatorias = ["{DataInicialRegistro}", "{DataFinalRegistro}", "{idMunicipio}", "{size}"]
                    tags_faltantes = [t for t in tags_obrigatorias if t not in url_consulta]
                    if tags_faltantes:
                        st.error(f"⚠️ Execução Abortada: O Endereço da Tela de Consulta cadastrado para este serviço SROP é inválido por não conter as tags obrigatórias: {', '.join(tags_faltantes)}. Por favor, corrija o cadastro do serviço.")
                        st.stop()
                        
                    import shutil, os
                    
                    st.info("💡 **A extração começou!** Se você perceber que colocou filtros errados ou quiser desistir, clique no botão **Cancelar Extração Agora** acima.")
                        
                    with st.status("Iniciando consulta e extração...", expanded=True) as status_ui:
                        def atualizar_status(mensagem: str):
                            status_ui.update(label=mensagem, state="running", expanded=True)
                            st.write(mensagem)
                        
                        # Pasta temporária para salvar os PDFs baixados
                        temp_dir = os.path.join(os.getcwd(), "dados_pdf_temp")
                        if os.path.exists(temp_dir):
                            shutil.rmtree(temp_dir, ignore_errors=True)
                        os.makedirs(temp_dir, exist_ok=True)
                        
                        try:
                            import automacao as aut
                            import extrair_bo as ex_bo
                            import importlib
                            importlib.reload(ex_bo)
                            
                            # Formata as datas para YYYY-MM-DD
                            data_ini_str = data_inicial.strftime("%Y-%m-%d")
                            data_fim_str = data_final.strftime("%Y-%m-%d")
                            
                            pdf_paths = aut.consultar_e_baixar_srop(
                                session_state=st.session_state[cookie_key],
                                url_consulta=url_consulta,
                                url_pdf_template=url_pdf,
                                data_inicial=data_ini_str,
                                data_final=data_fim_str,
                                temp_dir=temp_dir,
                                status_callback=atualizar_status,
                                id_municipio=id_municipio_val,
                                size=size_val,
                                numero_boletim=bo_val,
                                data_ini_fato=data_ini_fato_str,
                                data_fim_fato=data_fim_fato_str
                            )
                            
                            if not pdf_paths:
                                status_ui.update(label="Integração finalizada.", state="complete", expanded=True)
                                st.warning("⚠️ Nenhum Boletim de Ocorrência foi retornado para o período selecionado.")
                                st.session_state.pop(result_key, None)
                            else:
                                atualizar_status(f"Processando {len(pdf_paths)} arquivos PDF baixados...")
                                
                                # Carrega os mapeamentos e listas de bairros/municípios do banco de dados
                                db_mappings = {
                                    "mapa_upms": db.obter_mapeamento_upms(),
                                    "mapa_nomes_mun": db.obter_mapeamento_nomes_municipios(),
                                    "mapa_nomes_bai": db.obter_mapeamento_nomes_bairros(),
                                    "mapa_alternativos_bai": db.obter_mapeamento_alternativo_bairros(),
                                    "mapa_mun_todos": db.obter_municipios_com_bairro_todos_unico()
                                }
                                
                                lista_municipios, bairros_por_mun = ex_bo.carregar_dados_banco()
                                
                                resultados = []
                                for idx, pdf_path in enumerate(pdf_paths):
                                    filename = os.path.basename(pdf_path)
                                    atualizar_status(f"Analisando texto do BO {idx+1}/{len(pdf_paths)}: {filename}...")
                                    
                                    texto = ""
                                    try:
                                        texto = ex_bo.extrair_texto_pdf(pdf_path)
                                    except Exception as e:
                                        st.error(f"Erro ao extrair texto do PDF {filename}: {str(e)}")
                                        
                                    if not texto:
                                        res = {
                                            "ARQUIVO": filename,
                                            "BO_NUMERO": "Erro na leitura",
                                            "NARRATIVA": "Erro ao extrair texto do arquivo PDF.",
                                            "PROVIDENCIAS": "NI"
                                        }
                                    else:
                                        # Processa usando o analisador unificado do script extrair_bo
                                        layout_id = int(df_servicos_sidebar[df_servicos_sidebar["ID"] == id_servico].iloc[0].get("Layout_ID", 1))
                                        res = ex_bo.processar_texto_bo(
                                            texto,
                                            filename,
                                            layout_id,
                                            db_mappings,
                                            lista_municipios,
                                            bairros_por_mun
                                        )
                                    resultados.append(res)
                                    
                                # Compila tudo em um DataFrame e ordena
                                df_res = pd.DataFrame(resultados)
                                df_res = ex_bo.ordenar_dataframe(df_res, layout_id)
                                
                                # Ajusta o índice para iniciar em 1 para visualização
                                df_res.index = range(1, len(df_res) + 1)
                                
                                status_ui.update(label="Consulta e Extração concluídas com sucesso!", state="complete", expanded=True)
                                
                                # Guarda os resultados no session_state para persistir na tela após interações
                                st.session_state[result_key] = df_res
                                
                                # Comprime os PDFs baixados em um arquivo ZIP e guarda na sessão
                                import zipfile
                                import io
                                zip_buffer = io.BytesIO()
                                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                                    for pdf_path in pdf_paths:
                                        if os.path.exists(pdf_path):
                                            zip_file.write(pdf_path, os.path.basename(pdf_path))
                                
                                st.session_state[f"zip_{result_key}"] = zip_buffer.getvalue()
                                
                                st.success(f"🎉 Extração concluída! {len(resultados)} boletins processados com sucesso.")
                                
                        except Exception as error_exec:
                            status_ui.update(label="Ocorreu um erro na consulta.", state="error", expanded=True)
                            err_str = str(error_exec)
                            st.error(f"❌ {err_str}")
                            # Se a sessão expirou, removemos o estado da sessão para forçar novo login
                            if "Sessão expirada" in err_str:
                                st.session_state.pop(cookie_key, None)
                                db.limpar_sessao(id_servico)
                                st.warning("🔄 Sessão expirada no portal. O banco de dados foi limpo. Por favor, realize o login novamente.")
                                st.rerun()
                        finally:
                            # Limpa os PDFs temporários do servidor
                            if os.path.exists(temp_dir):
                                shutil.rmtree(temp_dir, ignore_errors=True)
                            # Desativa o botão de cancelar imediatamente após o término do processamento
                            placeholder_cancel.button("⏹️ Cancelar Processamento", use_container_width=True, disabled=True, key=f"btn_interrupt_done_{id_servico}")
                
                if btn_download_only:
                    # Validação em tempo de execução para garantir que as tags obrigatórias estão na URL base
                    tags_obrigatorias = ["{DataInicialRegistro}", "{DataFinalRegistro}", "{idMunicipio}", "{size}"]
                    tags_faltantes = [t for t in tags_obrigatorias if t not in url_consulta]
                    if tags_faltantes:
                        st.error(f"⚠️ Execução Abortada: O Endereço da Tela de Consulta cadastrado para este serviço SROP é inválido por não conter as tags obrigatórias: {', '.join(tags_faltantes)}. Por favor, corrija o cadastro do serviço.")
                        st.stop()
                        
                    import shutil, os
                    
                    st.info("💡 **O download dos PDFs começou!** Se quiser cancelar, clique no botão **Cancelar Processamento** acima.")
                        
                    with st.status("Iniciando consulta e download dos PDFs...", expanded=True) as status_ui:
                        def atualizar_status_dl(mensagem: str):
                            status_ui.update(label=mensagem, state="running", expanded=True)
                            st.write(mensagem)
                        
                        # Pasta temporária para salvar os PDFs baixados
                        temp_dir = os.path.join(os.getcwd(), "dados_pdf_temp")
                        if os.path.exists(temp_dir):
                            shutil.rmtree(temp_dir, ignore_errors=True)
                        os.makedirs(temp_dir, exist_ok=True)
                        
                        try:
                            import automacao as aut
                            
                            # Formata as datas para YYYY-MM-DD
                            data_ini_str = data_inicial.strftime("%Y-%m-%d")
                            data_fim_str = data_final.strftime("%Y-%m-%d")
                            
                            pdf_paths = aut.consultar_e_baixar_srop(
                                session_state=st.session_state[cookie_key],
                                url_consulta=url_consulta,
                                url_pdf_template=url_pdf,
                                data_inicial=data_ini_str,
                                data_final=data_fim_str,
                                temp_dir=temp_dir,
                                status_callback=atualizar_status_dl,
                                id_municipio=id_municipio_val,
                                size=size_val,
                                numero_boletim=bo_val,
                                data_ini_fato=data_ini_fato_str,
                                data_fim_fato=data_fim_fato_str
                            )
                            
                            if not pdf_paths:
                                status_ui.update(label="Integração finalizada.", state="complete", expanded=True)
                                st.warning("⚠️ Nenhum Boletim de Ocorrência foi retornado para o período selecionado.")
                                st.session_state.pop(f"zip_only_{id_servico}", None)
                            else:
                                atualizar_status_dl(f"Compactando {len(pdf_paths)} arquivos PDF baixados...")
                                
                                # Comprime os PDFs baixados em um arquivo ZIP e guarda na sessão
                                import zipfile
                                import io
                                zip_buffer = io.BytesIO()
                                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                                    for pdf_path in pdf_paths:
                                        if os.path.exists(pdf_path):
                                            zip_file.write(pdf_path, os.path.basename(pdf_path))
                                
                                st.session_state[f"zip_only_{id_servico}"] = zip_buffer.getvalue()
                                status_ui.update(label="Download concluído com sucesso!", state="complete", expanded=True)
                                st.success(f"🎉 Download concluído! {len(pdf_paths)} PDFs baixados com sucesso e disponíveis para download abaixo.")
                                # Limpa a chave do OCR caso ela existisse para não misturar resultados antigos
                                st.session_state.pop(result_key, None)
                                
                        except Exception as error_exec:
                            status_ui.update(label="Ocorreu um erro na consulta.", state="error", expanded=True)
                            err_str = str(error_exec)
                            st.error(f"❌ {err_str}")
                            if "Sessão expirada" in err_str:
                                st.session_state.pop(cookie_key, None)
                                db.limpar_sessao(id_servico)
                                st.warning("🔄 Sessão expirada no portal. O banco de dados foi limpo. Por favor, realize o login novamente.")
                                st.rerun()
                        finally:
                            # Limpa os PDFs temporários do servidor
                            if os.path.exists(temp_dir):
                                shutil.rmtree(temp_dir, ignore_errors=True)
                            # Desativa o botão de cancelar
                            placeholder_cancel.button("⏹️ Cancelar Processamento", use_container_width=True, disabled=True, key=f"btn_interrupt_dl_done_{id_servico}")
            
            # Exibe o botão de download do ZIP gerado por apenas baixar PDFs
            zip_only_key = f"zip_only_{id_servico}"
            if zip_only_key in st.session_state:
                st.markdown("---")
                st.subheader("📥 Arquivos PDF Disponíveis para Download")
                
                dt_ini_lbl = "consulta"
                dt_fim_lbl = "consulta"
                if f"exec_dt_ini_{id_servico}" in st.session_state:
                    dt_ini_lbl = st.session_state[f"exec_dt_ini_{id_servico}"].strftime('%d-%m-%Y')
                if f"exec_dt_fim_{id_servico}" in st.session_state:
                    dt_fim_lbl = st.session_state[f"exec_dt_fim_{id_servico}"].strftime('%d-%m-%Y')
                
                nome_zip = f"srop_bo_por_data_registro_{dt_ini_lbl}-{dt_fim_lbl}.zip"
                
                c_dl1, c_dl2 = st.columns([3, 1])
                with c_dl1:
                    st.download_button(
                        label="📥 Baixar PDFs Compactados (.zip)",
                        data=st.session_state[zip_only_key],
                        file_name=nome_zip,
                        mime="application/zip",
                        use_container_width=True,
                        key=f"btn_dl_zip_only_{id_servico}"
                    )
                with c_dl2:
                    if st.button("🗑️ Limpar Arquivos", key=f"btn_clear_zip_only_{id_servico}", use_container_width=True):
                        st.session_state.pop(zip_only_key, None)
                        st.rerun()
            
            # Exibe o resultado e o botão de download se houver dados no session_state
            if result_key in st.session_state:
                df_result = st.session_state[result_key]
                
                st.markdown("---")
                st.subheader("📋 Boletins Extraídos da Consulta")
                st.dataframe(df_result, width="stretch")
                
                # Gera o arquivo Excel para Download
                import io
                buffer_xlsx = io.BytesIO()
                with pd.ExcelWriter(buffer_xlsx, engine='openpyxl') as writer:
                    df_result.to_excel(writer, index=False, sheet_name='BOs_SROP')
                    
                dt_ini_lbl = "consulta"
                dt_fim_lbl = "consulta"
                if f"exec_dt_ini_{id_servico}" in st.session_state:
                    dt_ini_lbl = st.session_state[f"exec_dt_ini_{id_servico}"].strftime('%d-%m-%Y')
                if f"exec_dt_fim_{id_servico}" in st.session_state:
                    dt_fim_lbl = st.session_state[f"exec_dt_fim_{id_servico}"].strftime('%d-%m-%Y')
                
                nome_download = f"SROP_BOs_Extraidos_{dt_ini_lbl}_a_{dt_fim_lbl}.xlsx"
                
                col_btn1, col_btn2 = st.columns(2)
                
                with col_btn1:
                    st.download_button(
                        label="📥 Baixar Planilha de BOs (.xlsx)",
                        data=buffer_xlsx.getvalue(),
                        file_name=nome_download,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        key=f"btn_dl_srop_{id_servico}"
                    )
                    
                with col_btn2:
                    zip_key = f"zip_{result_key}"
                    if zip_key in st.session_state:
                        # O formato atualizado solicitado pelo usuário
                        nome_zip = f"srop_bo_por_data_registro_{dt_ini_lbl}-{dt_fim_lbl}.zip"
                        st.download_button(
                            label="🗂️ Baixar PDFs Analisados (.zip)",
                            data=st.session_state[zip_key],
                            file_name=nome_zip,
                            mime="application/zip",
                            use_container_width=True,
                            key=f"btn_dl_zip_{id_servico}"
                        )
